#!/usr/bin/env python3
"""Export every price_total worksheet into the app's normalized JSON cache."""

from __future__ import annotations

import hashlib
import json
import math
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SERVER_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = SERVER_ROOT / "price_tracking" / "price_summarized_optimized.xlsx"
OUTPUT = SERVER_ROOT / "data" / "price-tracking" / "cache.json"


CONFIGS = [
    {"id": "london-gold", "aliases": ["伦敦金现", "伦敦金", "London gold"], "product": "伦敦金现", "group": "市场资产", "unit": "美元/盎司", "sourceName": "汇率表伦敦金", "sourceUrl": "https://www.huilvbiao.com/gold/xau"},
    {"id": "btc-cme", "aliases": ["BTC.CME", "BTC CME"], "product": "BTC", "group": "市场资产", "unit": "美元", "sourceName": "Yahoo Finance", "sourceUrl": "https://finance.yahoo.com/quote/BTC-USD/", "valueColumn": 3},
    {"id": "eth-cme", "aliases": ["ETH.CME", "ETH CME"], "product": "ETH", "group": "市场资产", "unit": "美元", "sourceName": "Yahoo Finance", "sourceUrl": "https://finance.yahoo.com/quote/ETH-USD/", "valueColumn": 3},
    {"id": "brent-crude", "aliases": ["ICE布油", "原油", "布伦特原油", "Brent"], "product": "布伦特原油", "group": "能源", "unit": "美元/桶", "sourceName": "Yahoo Finance / TradingEconomics", "sourceUrl": "https://finance.yahoo.com/quote/BZ=F/"},
    {"id": "r32", "aliases": ["R32", "制冷剂R32"], "product": "R32", "group": "化工与制冷剂", "unit": "元/吨", "sourceName": "卓创资讯", "sourceUrl": "https://www.sci99.com/monitor-1572-0.html"},
    {"id": "tdi", "aliases": ["TDI"], "product": "TDI", "group": "化工与制冷剂", "unit": "元/吨", "sourceName": "卓创资讯", "sourceUrl": "https://www.sci99.com/monitor-375-0.html"},
    {"id": "polymeric-mdi", "aliases": ["聚合MDI", "聚合 MDI"], "product": "聚合MDI", "group": "化工与制冷剂", "unit": "元/吨", "sourceName": "卓创资讯", "sourceUrl": "https://www.sci99.com/monitor-384-0.html"},
    {"id": "pure-mdi", "aliases": ["MDI", "纯MDI", "纯 MDI"], "product": "纯MDI", "group": "化工与制冷剂", "unit": "元/吨", "sourceName": "卓创资讯", "sourceUrl": "https://www.sci99.com/monitor-94717214-1.html"},
    {"id": "lng", "aliases": ["液化天然气", "LNG"], "product": "全国LNG出厂价", "group": "能源", "unit": "元/吨", "sourceName": "上海石油天然气交易中心", "sourceUrl": "https://www.shpgx.com/html/qgjg.html"},
    {"id": "q5500", "aliases": ["Q5500动力煤", "Q5500"], "product": "Q5500动力煤", "group": "能源", "unit": "元/吨", "sourceName": "中国煤炭市场网CCTD", "sourceUrl": "https://www.cctd.com.cn/index.php?m=content&c=index&a=lists&catid=747"},
    {"id": "q5000", "aliases": ["Q5000动力煤", "Q5000"], "product": "Q5000动力煤", "group": "能源", "unit": "元/吨", "sourceName": "中国煤炭市场网CCTD", "sourceUrl": "https://www.cctd.com.cn/index.php?m=content&c=index&a=lists&catid=747"},
    {"id": "q4500", "aliases": ["Q4500动力煤", "Q4500"], "product": "Q4500动力煤", "group": "能源", "unit": "元/吨", "sourceName": "中国煤炭市场网CCTD", "sourceUrl": "https://www.cctd.com.cn/index.php?m=content&c=index&a=lists&catid=747"},
    {"id": "potassium-chloride", "aliases": ["氯化钾", "氯化钾粉(62%)"], "product": "氯化钾粉(62%)", "group": "农化", "unit": "元/吨", "sourceName": "生意社", "sourceUrl": "https://m1.100ppi.com/vane/759-%E6%B0%AF%E5%8C%96%E9%92%BE.html"},
    {"id": "cobalt", "aliases": ["钴"], "product": "钴（≥99.8%）", "group": "新能源材料", "unit": "元/吨", "sourceName": "生意社", "sourceUrl": "https://m1.100ppi.com/vane/602-%E9%92%B4.html"},
    {"id": "lme-copper", "aliases": ["LME铜", "LME 铜"], "product": "LME铜", "group": "基础金属", "unit": "美元/吨", "sourceName": "中国铝业网CNAL", "sourceUrl": "https://market.cnal.com/historical/lmeofficial.html"},
    {"id": "lme-aluminum", "aliases": ["LME铝", "LME原铝", "LME 铝"], "product": "LME铝", "group": "基础金属", "unit": "美元/吨", "sourceName": "中国铝业网CNAL", "sourceUrl": "https://market.cnal.com/historical/lmeofficial.html"},
    {"id": "electrolytic-aluminum", "aliases": ["铝", "电解铝"], "product": "电解铝", "group": "基础金属", "unit": "元/吨", "sourceName": "卓创资讯", "sourceUrl": "https://www.sci99.com/monitor-643-1.html"},
    {"id": "hafnium", "aliases": ["铪", "金属铪", "Hafnium"], "product": "金属铪", "group": "AI与电子材料", "unit": "美元/千克", "sourceName": "Strategic Metals Invest", "sourceUrl": "https://strategicmetalsinvest.com/hafnium-prices/"},
    {"id": "urea", "aliases": ["尿素"], "product": "尿素", "group": "农化", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "rebar", "aliases": ["螺纹钢"], "product": "螺纹钢", "group": "黑色与工业品", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "coking-coal", "aliases": ["焦煤"], "product": "焦煤", "group": "能源", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "alumina", "aliases": ["氧化铝"], "product": "氧化铝", "group": "基础金属", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "lithium-carbonate", "aliases": ["碳酸锂"], "product": "碳酸锂", "group": "新能源材料", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "coke", "aliases": ["焦炭"], "product": "焦炭", "group": "能源", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "polysilicon", "aliases": ["多晶硅"], "product": "多晶硅", "group": "新能源材料", "unit": "元/吨", "sourceName": "AKShare国内期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "comex-silver", "aliases": ["COMEX白银", "COMEX白银期货"], "product": "COMEX白银", "group": "新能源材料", "unit": "美元/盎司", "sourceName": "AKShare外盘期货", "sourceUrl": "https://akshare.akfamily.xyz/data/futures/futures.html"},
    {"id": "wolframite-65", "aliases": ["黑钨精矿", "黑钨精矿≧65%", "钨精矿"], "product": "黑钨精矿≧65%", "group": "AI与电子材料", "unit": "元/吨", "sourceName": "中钨在线", "sourceUrl": "https://www.ctia.com.cn/"},
    {"id": "waste-tungsten-bar", "aliases": ["废钨棒材"], "product": "废钨棒材", "group": "AI与电子材料", "unit": "元/千克", "sourceName": "中钨在线", "sourceUrl": "https://www.ctia.com.cn/"},
    {"id": "tungsten-powder", "aliases": ["钨粉"], "product": "钨粉", "group": "AI与电子材料", "unit": "元/千克", "sourceName": "中钨在线", "sourceUrl": "https://www.ctia.com.cn/"},
    {"id": "compound-fertilizer", "aliases": ["国产三元复合肥", "三元复合肥"], "product": "国产三元复合肥", "group": "农化", "unit": "元/吨", "sourceName": "price_total历史数据", "sourceUrl": ""},
    {"id": "zirconium-sponge", "aliases": ["海绵锆"], "product": "海绵锆", "group": "AI与电子材料", "unit": "元/千克", "sourceName": "price_total历史数据", "sourceUrl": ""},
    {"id": "chromium", "aliases": ["铬"], "product": "铬", "group": "基础金属", "unit": "元/吨", "sourceName": "price_total历史数据", "sourceUrl": ""},
    {"id": "tin", "aliases": ["锡"], "product": "锡", "group": "AI与电子材料", "unit": "元/吨", "sourceName": "price_total历史数据", "sourceUrl": ""},
]

GROUP_ORDER = {name: index for index, name in enumerate(["AI与电子材料", "基础金属", "新能源材料", "化工与制冷剂", "能源", "黑色与工业品", "农化", "市场资产", "其他"])}


def normalized(value: str) -> str:
    return re.sub(r"[\s（）()_.:/\\-]+", "", str(value or "")).lower()


def config_for_sheet(sheet_name: str) -> dict[str, Any]:
    target = normalized(sheet_name)
    exact_matches = []
    partial_matches = []
    for config in CONFIGS:
        for alias in config["aliases"]:
            alias_key = normalized(alias)
            if target == alias_key:
                exact_matches.append((len(alias_key), config))
            elif alias_key in target or target in alias_key:
                partial_matches.append((len(alias_key), config))
    if exact_matches:
        return max(exact_matches, key=lambda pair: pair[0])[1]
    if partial_matches:
        return max(partial_matches, key=lambda pair: pair[0])[1]
    clean_name = re.sub(r"(?:20)?\d{4}", "", sheet_name).strip(" （()）") or sheet_name
    digest = hashlib.sha1(sheet_name.encode("utf-8")).hexdigest()[:10]
    return {
        "id": f"workbook-{digest}",
        "aliases": [sheet_name, clean_name],
        "product": clean_name,
        "group": "其他",
        "unit": "",
        "sourceName": "price_total历史数据",
        "sourceUrl": "",
    }


def parse_date(cell: Any) -> date | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and getattr(cell, "is_date", False):
        try:
            return value.date() if isinstance(value, datetime) else value
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in (r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", r"(20\d{2})(\d{2})(\d{2})"):
            match = re.search(pattern, text)
            if match:
                try:
                    return date(*(int(part) for part in match.groups()))
                except ValueError:
                    return None
    return None


def parse_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if isinstance(value, str):
        match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", value.replace("，", ","))
        if match:
            try:
                return float(match.group(0).replace(",", ""))
            except ValueError:
                return None
    return None


def extract_history(worksheet: Any, config: dict[str, Any]) -> list[dict[str, Any]]:
    by_date: dict[str, float] = {}
    max_rows = min(worksheet.max_row or 0, 10000)
    max_cols = min(worksheet.max_column or 0, 10)
    for row in worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
        date_index = None
        parsed = None
        for index, cell in enumerate(row):
            parsed = parse_date(cell)
            if parsed:
                date_index = index
                break
        if parsed is None or date_index is None:
            continue
        price = None
        value_column = config.get("valueColumn")
        if isinstance(value_column, int) and 1 <= value_column <= len(row):
            price = parse_number(row[value_column - 1].value)
        else:
            for cell in row[date_index + 1:]:
                price = parse_number(cell.value)
                if price is not None:
                    break
        if price is None:
            continue
        by_date[parsed.isoformat()] = price
    return [{"date": key, "value": by_date[key]} for key in sorted(by_date)][-365:]


def normalize_history(config: dict[str, Any], history: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if config.get("id") == "wolframite-65":
        return [
            {**point, "value": point["value"] * 10000 if abs(point["value"]) < 1000 else point["value"]}
            for point in history
        ]
    return history


def business_days_old(latest_date: str | None) -> int:
    if not latest_date:
        return 9999
    current = datetime.strptime(latest_date, "%Y-%m-%d").date()
    today = date.today()
    count = 0
    while current < today:
        current += timedelta(days=1)
        if current.weekday() < 5:
            count += 1
    return count


def build_product(config: dict[str, Any], history: list[dict[str, Any]], sheet_name: str) -> dict[str, Any]:
    latest = history[-1] if history else None
    previous = history[-2] if len(history) > 1 else None
    move_percent = 0.0
    if latest and previous and previous["value"]:
        move_percent = (latest["value"] - previous["value"]) / previous["value"] * 100
    return {
        **config,
        "aliases": sorted(set(config.get("aliases", []) + [sheet_name, config["product"]])),
        "sheetNames": [sheet_name],
        "latestDate": latest["date"] if latest else None,
        "latestPrice": latest["value"] if latest else None,
        "movePercent": round(move_percent, 6),
        "stale": business_days_old(latest["date"] if latest else None) > 2,
        "history": history,
    }


def choose_duplicate(current: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    current_date = current.get("latestDate") or ""
    incoming_date = incoming.get("latestDate") or ""
    chosen = incoming if (incoming_date, len(incoming.get("history", []))) > (current_date, len(current.get("history", []))) else current
    chosen = dict(chosen)
    chosen["sheetNames"] = sorted(set(current.get("sheetNames", []) + incoming.get("sheetNames", [])))
    chosen["aliases"] = sorted(set(current.get("aliases", []) + incoming.get("aliases", [])))
    return chosen


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    products: dict[str, dict[str, Any]] = {}
    for sheet_name in workbook.sheetnames:
        if normalized(sheet_name) in {"总表", "summary"}:
            continue
        config = config_for_sheet(sheet_name)
        product = build_product(config, normalize_history(config, extract_history(workbook[sheet_name], config)), sheet_name)
        products[config["id"]] = choose_duplicate(products[config["id"]], product) if config["id"] in products else product

    generated_at = datetime.now().astimezone().isoformat()
    rows = list(products.values())
    if OUTPUT.exists():
        try:
            existing_rows = {item["id"]: item for item in json.loads(OUTPUT.read_text(encoding="utf-8")).get("products", [])}
            merged_rows = []
            for row in rows:
                existing = existing_rows.get(row["id"])
                if not existing:
                    merged_rows.append(row)
                    continue
                # Explicit value columns indicate a known worksheet schema.
                # Do not merge legacy cache points that may have been exported
                # from a different column (for example BTC/ETH volume instead
                # of close price).
                if row.get("valueColumn"):
                    merged_rows.append(row)
                    continue
                if set(existing.get("sheetNames", [])) != set(row.get("sheetNames", [])):
                    merged_rows.append(row)
                    continue
                history_by_date = {point["date"]: point for point in existing.get("history", [])}
                history_by_date.update({point["date"]: point for point in row.get("history", [])})
                merged_history = [history_by_date[key] for key in sorted(history_by_date)][-365:]
                preferred = existing if (existing.get("latestDate") or "") > (row.get("latestDate") or "") else row
                merged = {**row, **preferred, "history": merged_history}
                if merged_history:
                    merged["latestDate"] = merged_history[-1]["date"]
                    merged["latestPrice"] = merged_history[-1]["value"]
                    previous = merged_history[-2]["value"] if len(merged_history) > 1 else None
                    merged["movePercent"] = round((merged["latestPrice"] - previous) / previous * 100, 6) if previous else 0.0
                    merged["stale"] = business_days_old(merged["latestDate"]) > 2
                merged_rows.append(merged)
            rows = merged_rows
        except Exception:
            pass
    rows.sort(key=lambda item: (GROUP_ORDER.get(item.get("group", "其他"), 99), item.get("product", "")))
    payload = {
        "version": 1,
        "generatedAt": generated_at,
        "sourceWorkbook": "server/price_tracking/price_summarized_optimized.xlsx",
        "products": [{**row, "generatedAt": generated_at} for row in rows],
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    temp = OUTPUT.with_suffix(".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(OUTPUT)
    print(f"exported {len(rows)} price products to {OUTPUT}")


if __name__ == "__main__":
    main()
