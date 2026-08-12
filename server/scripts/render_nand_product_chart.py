#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render a single NAND spot-price product history chart from nand_spot.xlsx."""

from __future__ import annotations

import argparse
import re
from datetime import datetime, date
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties, fontManager
from matplotlib.ticker import FuncFormatter, MaxNLocator
from openpyxl import load_workbook


FONT_CANDIDATES = [
    "/System/Library/Fonts/PingFang.ttc",
    "/System/Library/Fonts/STHeiti Medium.ttc",
    "/System/Library/Fonts/Hiragino Sans GB.ttc",
    "/System/Library/Fonts/Songti.ttc",
    "/Library/Fonts/Songti.ttc",
    "/System/Library/Fonts/Supplemental/Songti.ttc",
]


def safe_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_.-]+", "_", value).strip("_") or "nand_product"


def setup_fonts():
    for font_path in FONT_CANDIDATES:
        path = Path(font_path)
        if not path.exists():
            continue
        try:
            fontManager.addfont(str(path))
            family = FontProperties(fname=str(path)).get_name()
            plt.rcParams["font.sans-serif"] = [family, "DejaVu Sans"]
            plt.rcParams["font.family"] = "sans-serif"
            plt.rcParams["axes.unicode_minus"] = False
            return FontProperties(fname=str(path))
        except Exception:
            continue
    plt.rcParams["axes.unicode_minus"] = False
    return None


def to_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    raise ValueError(f"unsupported date value: {value!r}")


def find_sheet(wb, product: str):
    if product in wb.sheetnames:
        return wb[product]
    normalized = product.replace("/", "_").strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == normalized:
            return wb[sheet_name]
    for sheet_name in wb.sheetnames:
        if sheet_name != "今日" and normalized in sheet_name.lower():
            return wb[sheet_name]
    raise RuntimeError(f"sheet not found for product: {product}")


def parse_history(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            day = to_datetime(row[0])
            price = float(row[5])
        except Exception:
            continue
        rows.append((day, price))
    rows.sort(key=lambda item: item[0])
    if not rows:
        raise RuntimeError("no valid history rows")
    return rows


def render(root: Path, product: str) -> Path:
    xlsx_path = root / "state" / "nand_spot.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    out_dir = root / "state" / "product_charts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{safe_name(product)}_history.png"

    cjk = setup_fonts()
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = find_sheet(wb, product)
    rows = parse_history(ws)
    dates = [item[0] for item in rows]
    prices = [item[1] for item in rows]

    fig, ax = plt.subplots(figsize=(9.6, 5.2), facecolor="#FCFCFD")
    ax.set_facecolor("#FFFFFF")
    color = "#1F4E78" if "64Gb" in product else "#2E75B6"

    ax.plot(dates, prices, color=color, linewidth=2.7, marker="o", markersize=6, zorder=3)
    ax.scatter(dates[-1], prices[-1], color=color, s=110, edgecolor="white", linewidth=1.5, zorder=4)

    y_min, y_max = min(prices), max(prices)
    y_span = max(y_max - y_min, 0.01)
    y_pad = max(y_span * 0.28, y_max * 0.015, 0.2)
    ax.set_ylim(max(0, y_min - y_pad), y_max + y_pad)
    ax.yaxis.set_major_locator(MaxNLocator(nbins=6))
    ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"${value:.2f}"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates) // 6)))

    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color("#D7DBE7")
    ax.grid(True, axis="y", color="#E7EAF2", linewidth=0.9)
    ax.grid(False, axis="x")
    ax.tick_params(axis="x", colors="#687086", labelsize=10)
    ax.tick_params(axis="y", colors="#1F2430", labelsize=10)

    for idx, (day, price) in enumerate(rows):
        if idx != len(rows) - 1 and len(rows) > 8:
            continue
        ax.annotate(
            f"${price:.2f}",
            xy=(day, price),
            xytext=(8 if idx == len(rows) - 1 else 0, 8),
            textcoords="offset points",
            color=color,
            fontsize=10,
            fontweight="bold",
            ha="left" if idx == len(rows) - 1 else "center",
        )

    if len(prices) >= 2 and prices[0]:
        interval_change = prices[-1] / prices[0] - 1
        subtitle = f"{dates[0].strftime('%Y-%m-%d')} 至 {dates[-1].strftime('%Y-%m-%d')} · 区间涨跌 {interval_change:+.1%}"
    else:
        subtitle = dates[-1].strftime("%Y-%m-%d")

    title_kwargs = {"fontsize": 16, "fontweight": "bold", "color": "#1F2430"}
    subtitle_kwargs = {"fontsize": 10, "color": "#687086"}
    if cjk is not None:
        title_kwargs["fontproperties"] = cjk
        subtitle_kwargs["fontproperties"] = cjk
    fig.text(0.08, 0.94, f"{product} · 历史价格走势", **title_kwargs)
    fig.text(0.08, 0.885, f"TrendForce · 盘平均 · 单位: 美元/颗 · {subtitle}", **subtitle_kwargs)

    ax.set_xlabel("日期", fontsize=11, labelpad=8, fontproperties=cjk)
    ax.set_ylabel("现货价", fontsize=11, labelpad=8, fontproperties=cjk)
    fig.subplots_adjust(left=0.10, right=0.94, top=0.80, bottom=0.16)
    fig.savefig(out_path, dpi=140, facecolor="#FCFCFD")
    plt.close(fig)
    print(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True)
    parser.add_argument("--product", required=True)
    args = parser.parse_args()
    render(Path(args.root), args.product)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
