from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


LIST_PAGE_URL = "https://www.ctia.com.cn/news/tungsten-news"
DEFAULT_WORKBOOK_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_PRICE_MAP = {
    "黑钨精矿": "black_tungsten_concentrate",
    "废钨棒材": "recycled_tungsten_bar",
    "钨粉": "tungsten_powder",
}
DEFAULT_WECHAT_FETCHER = Path.home() / "Documents" / "新闻资讯" / "scripts" / "fetch_wechat_mp.py"
WECHAT_FETCHER = Path(os.environ.get("CTIA_WECHAT_FETCHER", DEFAULT_WECHAT_FETCHER))
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
}
PRICE_PATTERNS = {
    "黑钨精矿": re.compile(r"65%黑钨精矿价格\s*([0-9]+(?:\.[0-9]+)?)\s*万元/标吨"),
    "废钨棒材": re.compile(r"废钨棒材价格\s*([0-9]+(?:\.[0-9]+)?)\s*元/千克"),
    "钨粉": re.compile(r"(?<!碳化)钨粉价格\s*([0-9]+(?:\.[0-9]+)?)\s*元/(?:千克|公斤)"),
}


@dataclass(frozen=True)
class PriceSnapshot:
    trade_date: date
    black_tungsten_concentrate: float
    recycled_tungsten_bar: float
    tungsten_powder: float
    source_url: str


def fetch_html(url: str, session: requests.Session | None = None) -> str:
    http = session or requests.Session()
    response = http.get(url, headers=HEADERS, timeout=30)
    response.raise_for_status()
    return response.text


def parse_article_date(article_node) -> date | None:
    time_node = article_node.select_one("time")
    if time_node is None:
        return None

    datetime_value = (time_node.get("datetime") or "").strip()
    if datetime_value:
        return date.fromisoformat(datetime_value[:10])

    text_value = time_node.get_text(" ", strip=True)
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text_value)
    if match is None:
        return None

    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def extract_today_article_links(list_page_html: str, target_date: date) -> list[str]:
    soup = BeautifulSoup(list_page_html, "html.parser")
    article_links: list[str] = []

    for article_node in soup.select("article.post"):
        article_date = parse_article_date(article_node)
        if article_date != target_date:
            continue

        link_node = article_node.select_one("div.content > a[href], a[itemprop='url'][href]")
        if link_node is None:
            continue

        href = link_node.get("href", "").strip()
        if not href:
            continue

        article_links.append(urljoin(LIST_PAGE_URL, href))

    return article_links


def extract_article_text(article_html: str) -> str:
    soup = BeautifulSoup(article_html, "html.parser")
    content_node = soup.select_one("#js_content")
    if content_node is None:
        content_node = soup.select_one(".rich_media_content")
    if content_node is None:
        content_node = soup.select_one(".post-content")
    if content_node is None:
        content_node = soup.select_one("article")
    if content_node is None:
        raise ValueError("Could not find the CTIA article content block.")

    return content_node.get_text("\n", strip=True)


def extract_prices(article_text: str) -> dict[str, float]:
    extracted: dict[str, float] = {}
    normalized_text = article_text.replace("\xa0", " ")

    for sheet_name, pattern in PRICE_PATTERNS.items():
        match = pattern.search(normalized_text)
        if match is not None:
            extracted[sheet_name] = float(match.group(1))

    return extracted


def build_snapshot(target_date: date, article_url: str, article_text: str) -> PriceSnapshot:
    parsed_prices = extract_prices(article_text)
    missing = [name for name in SHEET_PRICE_MAP if name not in parsed_prices]
    if missing:
        raise ValueError(f"Missing tungsten prices in article: {', '.join(missing)}")

    return PriceSnapshot(
        trade_date=target_date,
        black_tungsten_concentrate=parsed_prices["黑钨精矿"],
        recycled_tungsten_bar=parsed_prices["废钨棒材"],
        tungsten_powder=parsed_prices["钨粉"],
        source_url=article_url,
    )


def parse_wechat_articles(stdout: str) -> list[dict[str, Any]]:
    match = re.search(r"(\[\s*\{[\s\S]*\}\s*\])\s*$", stdout)
    if match is None:
        raise ValueError("WeChat fetcher returned no JSON article list.")
    payload = json.loads(match.group(1))
    return payload if isinstance(payload, list) else []


def fetch_wechat_snapshot(target_date: date, session: requests.Session | None = None) -> PriceSnapshot:
    if not WECHAT_FETCHER.exists():
        raise FileNotFoundError(f"WeChat fetcher not found: {WECHAT_FETCHER}")

    result = subprocess.run(
        [
            sys.executable,
            str(WECHAT_FETCHER),
            "--source",
            "chinatungsten",
            "--since",
            "72h",
            "--limit",
            "5",
        ],
        capture_output=True,
        text=True,
        timeout=60,
        check=False,
    )
    if result.returncode != 0:
        message = (result.stderr or result.stdout or "WeChat fetch failed").strip()
        raise RuntimeError(message[-800:])

    articles = parse_wechat_articles(result.stdout)
    for article in articles:
        published_raw = str(article.get("published_at") or "")
        try:
            published_date = datetime.fromisoformat(published_raw.replace("Z", "+00:00")).astimezone().date()
        except ValueError:
            published_date = None
        title = str(article.get("title") or "")
        if published_date != target_date and target_date.isoformat() not in title:
            continue
        article_url = str(article.get("url") or "")
        if not article_url:
            continue
        http = session or requests.Session()
        response = http.get(
            article_url,
            headers={**HEADERS, "Referer": "https://mp.weixin.qq.com/"},
            timeout=30,
        )
        response.raise_for_status()
        return build_snapshot(target_date, article_url, extract_article_text(response.text))

    raise ValueError(f"No 中钨在线 WeChat article was found for {target_date.isoformat()}.")


def fetch_website_snapshot(target_date: date, session: requests.Session | None = None) -> PriceSnapshot:
    list_page_html = fetch_html(LIST_PAGE_URL, session=session)
    article_links = extract_today_article_links(list_page_html, target_date)
    if not article_links:
        raise ValueError(f"No CTIA tungsten news articles were found for {target_date.isoformat()}.")

    for article_url in article_links:
        article_html = fetch_html(article_url, session=session)
        article_text = extract_article_text(article_html)
        try:
            return build_snapshot(target_date, article_url, article_text)
        except ValueError:
            continue

    raise ValueError(f"Could not find all tracked tungsten prices in CTIA articles for {target_date.isoformat()}.")


def fetch_today_snapshot(target_date: date, session: requests.Session | None = None) -> PriceSnapshot:
    errors: list[str] = []
    try:
        return fetch_wechat_snapshot(target_date, session=session)
    except Exception as error:
        errors.append(f"WeChat: {error}")
    try:
        return fetch_website_snapshot(target_date, session=session)
    except Exception as error:
        errors.append(f"Website: {error}")
    raise ValueError(" | ".join(errors))


def normalize_trade_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def clone_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def ensure_row_format(worksheet: Worksheet, row_index: int) -> None:
    template_row = row_index + 1 if worksheet.max_row >= row_index + 1 else None
    if template_row is not None:
        for column in (1, 2):
            clone_cell_style(worksheet.cell(row=template_row, column=column), worksheet.cell(row=row_index, column=column))

    worksheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"
    worksheet.cell(row=row_index, column=2).number_format = "#,##0.00_ "


def upsert_sheet_row(worksheet: Worksheet, trade_date: date, price: float) -> str:
    for row_index in range(3, worksheet.max_row + 1):
        existing_date = normalize_trade_date(worksheet.cell(row=row_index, column=1).value)
        if existing_date == trade_date:
            worksheet.cell(row=row_index, column=1, value=trade_date)
            worksheet.cell(row=row_index, column=2, value=price)
            ensure_row_format(worksheet, row_index)
            return "updated"

    worksheet.insert_rows(3)
    ensure_row_format(worksheet, 3)
    worksheet.cell(row=3, column=1, value=trade_date)
    worksheet.cell(row=3, column=2, value=price)
    return "inserted"


def update_workbook(workbook_path: Path, snapshot: PriceSnapshot) -> dict[str, str]:
    workbook = load_workbook(workbook_path)
    actions: dict[str, str] = {}

    for sheet_name, attribute_name in SHEET_PRICE_MAP.items():
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.cell(row=2, column=1, value="指标名称")
            worksheet.cell(row=2, column=2, value=f"中国:报价:{sheet_name}")
        else:
            worksheet = workbook[sheet_name]

        price_value = float(getattr(snapshot, attribute_name))
        actions[sheet_name] = upsert_sheet_row(worksheet, snapshot.trade_date, price_value)

    workbook.save(workbook_path)
    return actions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch CTIA tungsten prices and update the workbook.")
    parser.add_argument(
        "--date",
        dest="trade_date",
        default=date.today().isoformat(),
        help="Target trade date in YYYY-MM-DD format. Defaults to today.",
    )
    parser.add_argument(
        "--workbook",
        default=DEFAULT_WORKBOOK_FILE,
        help="Workbook filename or absolute path. Defaults to price_summarized_optimized.xlsx.",
    )
    return parser.parse_args()


def resolve_workbook_path(workbook_arg: str) -> Path:
    workbook_path = Path(workbook_arg)
    if not workbook_path.is_absolute():
        workbook_path = Path(__file__).resolve().parent / workbook_path
    return workbook_path


def main() -> None:
    args = parse_args()
    target_date = date.fromisoformat(args.trade_date)
    workbook_path = resolve_workbook_path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    snapshot = fetch_today_snapshot(target_date)
    actions = update_workbook(workbook_path, snapshot)

    print(f"Target date: {snapshot.trade_date.isoformat()}")
    print(f"Source article: {snapshot.source_url}")
    print(f"黑钨精矿: {snapshot.black_tungsten_concentrate} 万元/标吨 ({actions['黑钨精矿']})")
    print(f"废钨棒材: {snapshot.recycled_tungsten_bar} 元/千克 ({actions['废钨棒材']})")
    print(f"钨粉: {snapshot.tungsten_powder} 元/千克 ({actions['钨粉']})")
    print(f"Updated workbook: {workbook_path}")


if __name__ == "__main__":
    main()
