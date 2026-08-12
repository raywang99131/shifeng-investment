"""100ppi 钴基准价 -> price_summarized_optimized.xlsx 钴工作表。

运行后会：
1. 从生意社（100ppi）抓取钴基准价走势；
2. 读取 price_summarized_optimized.xlsx 中“钴”工作表现有记录；
3. 按日期合并，按日期降序重写整张工作表，使首行（最新交易日）与抓取到的数据保持一致；
4. 工作表首两行格式与价格总表（row 7）口径保持一致：A2 为“指标名称”，B2 为指示标签。
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.datetime import from_excel


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "钴"
HEADER_LABEL = "指标名称"
INDICATOR_LABEL = "中国:市场价:钴(≥99.8%)"

DEFAULT_LOOKBACK_DAYS = 30
HISTORY_LOOKBACK_DAYS = 365
REQUEST_TIMEOUT = 20


@dataclass(frozen=True)
class SheetConfig:
    sheet_name: str
    indicator_label: str


CONFIG = SheetConfig(
    sheet_name=SHEET_NAME,
    indicator_label=INDICATOR_LABEL,
)


# -----------------------------------------------------------------------
# 100ppi 抓取
# -----------------------------------------------------------------------

def parse_mmdd_to_date(mmdd: str):
    """把 07-11 转为当前年份日期；如果解析出来是未来日期，则自动回退一年。"""
    mmdd = str(mmdd).strip()
    m = re.match(r"^(\d{1,2})-(\d{1,2})$", mmdd)
    if not m:
        return pd.NaT

    year = datetime.now().year
    month = int(m.group(1))
    day = int(m.group(2))

    dt = pd.Timestamp(year=year, month=month, day=day)
    today = pd.Timestamp.today().normalize()

    if dt > today + pd.Timedelta(days=3):
        dt = pd.Timestamp(year=year - 1, month=month, day=day)

    return dt


def fetch_cobalt_from_100ppi() -> pd.DataFrame:
    """从生意社抓取钴基准价走势，单位：元/吨。

    页面：
    https://m1.100ppi.com/vane/602-%E9%92%B4.html
    """
    commodity = quote("钴")
    url = f"https://m1.100ppi.com/vane/602-{commodity}.html"

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
    }

    resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding

    soup = BeautifulSoup(resp.text, "lxml")
    text = soup.get_text("\n", strip=True)

    # 匹配：
    # 07-11 383800.00 0.00%
    # 07-10 383800.00 0.03%
    # 跳过：
    # 07-12 - -%
    pattern = re.compile(
        r"(?P<date>\d{1,2}-\d{1,2})\s+"
        r"(?P<price>\d+(?:\.\d+)?)\s+"
        r"(?P<change>[+-]?\d+(?:\.\d+)?)%"
    )

    rows = []
    for m in pattern.finditer(text):
        rows.append({
            "date": parse_mmdd_to_date(m.group("date")),
            "price": float(m.group("price")),
            "daily_change_pct": float(m.group("change")) / 100,
        })

    if not rows:
        with open("debug_100ppi_cobalt.html", "w", encoding="utf-8") as f:
            f.write(resp.text)
        raise ValueError(
            "未从生意社页面解析到钴价格。"
            "已保存 debug_100ppi_cobalt.html，请检查是否被反爬或页面结构变化。"
        )

    df = pd.DataFrame(rows)
    df = df.dropna(subset=["date", "price"])
    df = df.drop_duplicates(subset=["date"], keep="first")
    df = df.sort_values("date").reset_index(drop=True)
    return df


def get_latest_cobalt_price() -> dict[str, Any]:
    """透出最近一天及前一天的价格概况，仅用于调试/CLI 查看。"""
    df = fetch_cobalt_from_100ppi().dropna(subset=["price"]).sort_values("date")

    latest = df.iloc[-1]
    prev = df.iloc[-2] if len(df) >= 2 else None

    result = {
        "date": latest["date"].date(),
        "today_price": float(latest["price"]),
        "unit": "元/吨",
        "source": "生意社-钴基准价",
    }
    if prev is not None:
        result["yesterday_price"] = float(prev["price"])
        result["change"] = float(latest["price"] - prev["price"])
        result["pct_change"] = float((latest["price"] - prev["price"]) / prev["price"])
    return result


# -----------------------------------------------------------------------
# 工作表合并 / 写入
# -----------------------------------------------------------------------

def normalize_trade_date(value: Any) -> date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        try:
            return pd.to_datetime(text).date()
        except Exception:
            return None
    return None


def normalize_price(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_existing_rows(worksheet) -> list[list[Any]]:
    rows: list[list[Any]] = []
    if worksheet is None:
        return rows
    for row in worksheet.iter_rows(min_row=3, max_col=2, values_only=True):
        trade_date = normalize_trade_date(row[0])
        price = normalize_price(row[1])
        if trade_date is None or price is None:
            continue
        rows.append([trade_date, price])
    return rows


def detect_fetch_start_date(worksheet) -> date:
    today = date.today()
    existing_rows = read_existing_rows(worksheet)
    if not existing_rows:
        return today - timedelta(days=HISTORY_LOOKBACK_DAYS)
    latest_existing = max(row[0] for row in existing_rows)
    return latest_existing - timedelta(days=DEFAULT_LOOKBACK_DAYS)


def fetch_cobalt_rows(start_date: date, end_date: date) -> list[list[Any]]:
    df = fetch_cobalt_from_100ppi()
    if df.empty:
        return []
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    df = df[(df["date"] >= start_ts) & (df["date"] <= end_ts)]
    out = []
    for _, row in df.iterrows():
        ts = row["date"]
        out.append([ts.date() if hasattr(ts, "date") else ts, float(row["price"])])
    return out


def merge_rows(existing_rows, new_rows):
    rows_by_date = {row[0]: row for row in existing_rows}
    inserted = 0
    updated = 0
    for row in new_rows:
        if row[0] in rows_by_date:
            updated += 1
        else:
            inserted += 1
        rows_by_date[row[0]] = row
    merged_rows = sorted(rows_by_date.values(), key=lambda item: item[0], reverse=True)
    return merged_rows, inserted, updated


def ensure_sheet(workbook):
    if CONFIG.sheet_name in workbook.sheetnames:
        return workbook[CONFIG.sheet_name]
    return workbook.create_sheet(CONFIG.sheet_name)


def rewrite_sheet(worksheet, rows: list[list[Any]]) -> None:
    if worksheet.max_row > 0:
        worksheet.delete_rows(1, worksheet.max_row)

    worksheet["A1"] = None
    worksheet["B1"] = None
    worksheet["A2"] = HEADER_LABEL
    worksheet["B2"] = CONFIG.indicator_label

    times_font = Font(name="Times New Roman")
    worksheet["A2"].font = times_font
    worksheet["B2"].font = times_font

    for trade_date, price in rows:
        worksheet.append([trade_date, price])

    for row in worksheet.iter_rows(min_row=3, max_col=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = times_font

    for cell in worksheet["A"][2:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in worksheet["B"][2:]:
        cell.number_format = "0.00"

    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 38


def save_workbook(workbook, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError:
        temp_path = workbook_path.with_name(f"{workbook_path.stem}.tmp{workbook_path.suffix}")
        workbook.save(temp_path)
        temp_path.replace(workbook_path)


# -----------------------------------------------------------------------
# 主流程
# -----------------------------------------------------------------------

def update_co_workbook(
    workbook_path: Path = WORKBOOK_PATH,
    end_date: date | None = None,
) -> dict[str, Any]:
    """读取现有的钴工作表 → 抓取缺失日期 → 按日期合并 → 重写整张工作表。"""
    if end_date is None:
        end_date = date.today()

    workbook = load_workbook(workbook_path)
    existing_sheet = (
        workbook[CONFIG.sheet_name]
        if CONFIG.sheet_name in workbook.sheetnames
        else None
    )
    start_date = detect_fetch_start_date(existing_sheet)
    new_rows = fetch_cobalt_rows(start_date=start_date, end_date=end_date)
    existing_rows = read_existing_rows(existing_sheet)
    merged_rows, inserted, updated = merge_rows(existing_rows, new_rows)

    worksheet = ensure_sheet(workbook)
    rewrite_sheet(worksheet, merged_rows)

    save_workbook(workbook, workbook_path)

    return {
        "workbook_path": str(workbook_path.resolve()),
        "total_rows": len(merged_rows),
        "inserted": inserted,
        "updated": updated,
        "fetch_start": start_date.isoformat(),
        "fetch_end": end_date.isoformat(),
    }


def main() -> None:
    summary = update_co_workbook()
    print(f"更新工作簿：{summary['workbook_path']}")
    print(
        f"{CONFIG.sheet_name}: total={summary['total_rows']}, "
        f"inserted={summary['inserted']}, updated={summary['updated']}, "
        f"fetch_window={summary['fetch_start']}~{summary['fetch_end']}"
    )


if __name__ == "__main__":
    main()
