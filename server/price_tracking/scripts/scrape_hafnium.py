from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PAGE_URL = "https://strategicmetalsinvest.com/hafnium-prices/"
WORKBOOK_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "\u94ea"
PREVIOUS_SHEET_NAME = "\u5236\u51b7\u5242R32"
HEADERS = [
    "Date",
    "Hafnium price (USD/kg)",
    "Change to today",
]


def fetch_page_html() -> str:
    response = requests.get(
        PAGE_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            ),
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.text


def parse_price(value: str) -> float:
    normalized = value.replace("$", "").replace(",", "").replace("/kg", "").strip()
    return float(normalized)


def parse_change(value: str) -> float | str:
    normalized = value.strip().replace("\xa0", " ")
    if normalized in {"", "-", "\u2014", "&mdash;"}:
        return "-"

    sign = -1 if normalized.startswith("-") else 1
    number = normalized.replace("%", "").replace("+", "").replace("-", "").replace(",", "").strip()
    return sign * float(number) / 100


def build_rows(html: str) -> list[list[Any]]:
    soup = BeautifulSoup(html, "html.parser")
    container = soup.select_one("div.metal-price-performance table tbody")
    if container is None:
        raise ValueError("Could not find the hafnium performance table on the page.")

    rows: list[list[Any]] = []
    for tr in container.find_all("tr"):
        cells = [cell.get_text(" ", strip=True) for cell in tr.find_all("td")]
        if len(cells) != 3:
            continue

        period, price_text, change_text = cells
        rows.append([period, parse_price(price_text), parse_change(change_text)])

    if not rows:
        raise ValueError("The hafnium performance table was found, but no rows were parsed.")

    return rows


def parse_page_trade_date(html: str) -> date:
    soup = BeautifulSoup(html, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    match = re.search(r"Price as of\s+([A-Za-z]{3,9}\s+\d{1,2}\s+\d{4})\b", page_text, re.IGNORECASE)
    if match is None:
        raise ValueError("Could not find the `Price as of ...` date on the hafnium page.")

    date_text = match.group(1)
    for fmt in ("%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(date_text, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Unsupported hafnium page date format: {date_text}")


def normalize_trade_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text)
        except ValueError:
            return None
    return None


def build_daily_row(html: str, trade_date: date) -> list[Any]:
    for row in build_rows(html):
        period = str(row[0]).strip().lower()
        if period == "current price":
            return [trade_date, row[1], row[2]]

    raise ValueError("Could not find the `Current price` row in the hafnium performance table.")


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="0B67D4")
    header_font = Font(name="Times New Roman", color="FFFFFF", bold=True)   # 修改标题字体
    body_font = Font(name="Times New Roman")                                # 正文字体
    alt_fill = PatternFill("solid", fgColor="EEF3F6")
    center = Alignment(horizontal="center", vertical="center")

    # 标题行
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 数据行（包括交替底色和字体）
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_index % 2 == 0 else None
        for cell in row:
            cell.alignment = center
            cell.font = body_font          # 设置正文字体
            if fill is not None:
                cell.fill = fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["C"].width = 18


def apply_number_formats(worksheet, rows: list[list[Any]]) -> None:
    for row_index, row in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_index, column=2).number_format = '$#,##0.00"/kg"'
        if isinstance(row[2], (int, float)):
            worksheet.cell(row=row_index, column=3).number_format = "+0.00%;-0.00%"


def load_existing_daily_rows(workbook_path: Path) -> list[list[Any]]:
    workbook = load_workbook(workbook_path)
    if SHEET_NAME not in workbook.sheetnames:
        return []

    worksheet = workbook[SHEET_NAME]
    rows: list[list[Any]] = []
    for row_index in range(2, worksheet.max_row + 1):
        trade_date = normalize_trade_date(worksheet.cell(row=row_index, column=1).value)
        if trade_date is None:
            continue

        price_value = worksheet.cell(row=row_index, column=2).value
        change_value = worksheet.cell(row=row_index, column=3).value
        if price_value in {None, ""}:
            continue

        rows.append([trade_date, float(price_value), change_value])

    return rows


def write_workbook(rows: list[list[Any]], workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    sheet_index = len(workbook.sheetnames)
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    if PREVIOUS_SHEET_NAME in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(PREVIOUS_SHEET_NAME) + 1

    worksheet = workbook.create_sheet(title=SHEET_NAME, index=sheet_index)
    worksheet.append(HEADERS)

    for row in rows:
        worksheet.append(row)

    style_sheet(worksheet)
    apply_number_formats(worksheet, rows)

    try:
        workbook.save(workbook_path)
    except PermissionError:
        buffer = BytesIO()
        workbook.save(buffer)
        with workbook_path.open("rb+") as handle:
            handle.seek(0)
            handle.write(buffer.getvalue())
            handle.truncate()


def upsert_daily_row(existing_rows: list[list[Any]], new_row: list[Any]) -> tuple[list[list[Any]], str]:
    trade_date = new_row[0]
    action = "inserted"
    updated_rows: list[list[Any]] = []

    for row in existing_rows:
        if row[1] == new_row[1] and row[2] == new_row[2] and row[0] > trade_date:
            continue
        if row[0] == trade_date:
            updated_rows.append(new_row)
            action = "updated"
        else:
            updated_rows.append(row)

    if action == "inserted":
        updated_rows.append(new_row)

    updated_rows.sort(key=lambda row: row[0], reverse=True)
    return updated_rows, action


def main() -> None:
    workbook_path = Path(__file__).resolve().parent / WORKBOOK_FILE
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    html = fetch_page_html()
    trade_date = parse_page_trade_date(html)
    current_row = build_daily_row(html, trade_date)
    existing_rows = load_existing_daily_rows(workbook_path)
    rows, action = upsert_daily_row(existing_rows, current_row)
    write_workbook(rows, workbook_path)

    print(f"Updated workbook: {workbook_path}")
    print(f"Inserted sheet: {SHEET_NAME}")
    print(f"Trade date: {trade_date.isoformat()} ({action})")
    print(f"Rows written: {len(rows)}")
    print(f"Hafnium price: ${current_row[1]:,.2f}/kg")


if __name__ == "__main__":
    main()
