# -*- coding: utf-8 -*-
"""
用 AKShare 替代工作簿中的国内期货型 Wind 数据。国内期货型标的为：
1. 尿素2609（氮肥） -> UR2609
2. 螺纹钢2610       -> RB2610
3. 焦煤2609         -> JM2609
4. 氧化铝2609       -> AO2609
5. 碳酸锂2609       -> LC2609
6. 焦炭2609         -> J2609
7. 多晶硅2609       -> PS2609
8. COMEX白银期货    -> SI
"""

from __future__ import annotations

import argparse
import copy
import json
import logging
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, to_excel


class DirectSinaMarketData:
    """Small AKShare-compatible fallback that avoids curl_cffi at import time."""

    @staticmethod
    def _get_jsonp(url: str, params: dict[str, str]) -> list:
        response = requests.get(
            url,
            params=params,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": "https://finance.sina.com.cn/futures/",
            },
            timeout=8,
        )
        response.raise_for_status()
        text = response.text
        start = text.find("[")
        end = text.rfind("]")
        if start < 0 or end < start:
            raise ValueError("Sina returned invalid JSONP")
        return json.loads(text[start : end + 1])

    def futures_zh_minute_sina(self, symbol: str, period: str = "1") -> pd.DataFrame:
        data = self._get_jsonp(
            "https://stock2.finance.sina.com.cn/futures/api/jsonp.php/=/InnerFuturesNewService.getFewMinLine",
            {"symbol": symbol, "type": period},
        )
        frame = pd.DataFrame(data)
        frame.columns = ["datetime", "open", "high", "low", "close", "volume", "hold"]
        return frame

    def futures_zh_daily_sina(self, symbol: str) -> pd.DataFrame:
        data = self._get_jsonp(
            "https://stock2.finance.sina.com.cn/futures/api/jsonp.php/var%20_V21052021_4_12=/InnerFuturesNewService.getDailyKLine",
            {"symbol": symbol, "type": "2021_04_12"},
        )
        frame = pd.DataFrame(data)
        frame.columns = ["date", "open", "high", "low", "close", "volume", "hold", "settle"]
        return frame

    def futures_foreign_hist(self, symbol: str) -> pd.DataFrame:
        today = datetime.today()
        stamp = f"{today.year}_{today.month}_{today.day}"
        data = self._get_jsonp(
            f"https://stock2.finance.sina.com.cn/futures/api/jsonp.php/var%20_S{stamp}=/GlobalFuturesService.getGlobalFuturesDailyKLine",
            {"symbol": symbol, "_": stamp, "source": "web"},
        )
        return pd.DataFrame(data)


ak = DirectSinaMarketData()

import sys
from pathlib import Path

# Ensure sibling-script imports work whether this file is run as
# ``python scripts/update_akshare_futures.py`` or imported as
# ``scripts.update_akshare_futures`` (e.g. by the pytest suite).
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

# ========== 1. 根据图片调整后的国内期货型品种映射 ==========
# price_col:
# - close  ：收盘价
# - settle ：结算价
#
# 当前先统一用 close。
# 如果你想改成结算价，把对应品种的 price_col 改成 "settle" 即可。
FUTURES_MAP = {
    "尿素2609（氮肥）": {
        "sheet_candidates": ["尿素2609（爱肥）", "尿素2609", "尿素"],
        "symbols": ["UR2609", "ur2609", "UR0", "ur0"],
        "price_col": "close",
        "desc": "期货收盘价:尿素2609；取不到则回退尿素连续",
    },
    "螺纹钢2610": {
        "sheet_candidates": ["螺纹钢2610", "螺纹钢"],
        "symbols": ["RB2610", "rb2610", "RB0", "rb0"],
        "price_col": "close",
        "desc": "期货收盘价:螺纹钢2610；取不到则回退螺纹钢连续",
    },
    "焦煤2609": {
        "sheet_candidates": ["焦煤2609", "焦煤"],
        "symbols": ["JM2609", "jm2609", "JM0", "jm0"],
        "price_col": "close",
        "desc": "期货收盘价:焦煤2609；取不到则回退焦煤连续",
    },
    "氧化铝2609": {
        "sheet_candidates": ["氧化铝2609", "氧化铝"],
        "symbols": ["AO2609", "ao2609", "AO0", "ao0"],
        "price_col": "close",
        "desc": "期货收盘价:氧化铝2609；取不到则回退氧化铝连续",
    },
    "碳酸锂2609": {
        "sheet_candidates": ["碳酸锂2609", "碳酸锂"],
        "symbols": ["LC2609", "lc2609", "LC0", "lc0"],
        "price_col": "close",
        "desc": "期货收盘价:碳酸锂2609；取不到则回退碳酸锂连续",
    },
    "焦炭2609": {
        "sheet_candidates": ["焦炭2609", "焦炭"],
        "symbols": ["J2609", "j2609", "J0", "j0"],
        "price_col": "close",
        "desc": "期货收盘价:焦炭2609；取不到则回退焦炭连续",
    },
    "多晶硅2609": {
        "sheet_candidates": ["多晶硅2609", "多晶硅"],
        "symbols": ["PS2609", "ps2609", "PS0", "ps0"],
        "price_col": "close",
        "desc": "期货收盘价:多晶硅2609；取不到则回退多晶硅连续",
    },
}


# ========== 1.1 外盘期货品种映射 ==========
# 使用 AKShare 的 futures_foreign_hist 接口。
# COMEX白银在新浪外盘代码表中对应 SI。
FOREIGN_FUTURES_MAP = {
    "COMEX白银期货": {
        "sheet_candidates": ["COMEX白银期货", "COMEX白银", "COMEX 白银", "白银COMEX", "白银期货COMEX"],
        "default_sheet_name": "COMEX白银期货",
        "symbols": ["SI"],
        "price_col": "close",
        "desc": "外盘期货收盘价:COMEX白银期货，单位:美元/盎司，数据源:AKShare futures_foreign_hist(SI)",
    },
}


@dataclass
class FetchResult:
    item_name: str
    sheet_name: str
    symbol: str
    trade_date: date
    price: float
    price_col: str
    action: str
    deleted_duplicates: int
    desc: str


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """统一列名，避免大小写、空格导致取列失败。"""
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def resolve_sheet_name(wb, sheet_candidates: list[str]) -> str | None:
    """
    从候选 sheet 名里找实际存在的 sheet。

    例如：
    图片里叫“尿素2609（爱肥）”，
    但工作簿 sheet 可能只叫“尿素”。
    """
    for name in sheet_candidates:
        if name in wb.sheetnames:
            return name
    return None


def resolve_or_create_sheet(
    wb,
    sheet_candidates: list[str],
    default_sheet_name: str,
) -> tuple[str, bool]:
    """
    优先使用已有 sheet；如果不存在，则创建一个新的 sheet。

    外盘新增品种可能原工作簿没有对应 sheet，因此不能像国内期货一样直接跳过。
    """
    sheet_name = resolve_sheet_name(wb, sheet_candidates)
    if sheet_name is not None:
        return sheet_name, False

    candidate = default_sheet_name
    if candidate in wb.sheetnames:
        i = 2
        while f"{default_sheet_name}_{i}" in wb.sheetnames:
            i += 1
        candidate = f"{default_sheet_name}_{i}"

    wb.create_sheet(candidate)
    return candidate, True


def fetch_latest_foreign_futures_daily(
    item_name: str,
    symbols: Iterable[str],
    price_col: str = "close",
) -> tuple[date, float, str, str]:
    """
    从 AKShare 拉取外盘期货最新日频数据。

    COMEX白银期货使用 symbol="SI"。
    返回字段通常包含 date/open/high/low/close/volume/position 等。
    """
    errors = []

    for symbol in dict.fromkeys(item.upper() for item in symbols):
        try:
            df = ak.futures_foreign_hist(symbol=symbol)

            if df is None or df.empty:
                raise ValueError("AKShare 返回空表")

            df = normalize_columns(df)

            if "date" not in df.columns:
                raise ValueError(f"缺少 date 列，实际列={list(df.columns)}")

            if price_col not in df.columns:
                raise ValueError(f"缺少 {price_col} 列，实际列={list(df.columns)}")

            df["_raw_order"] = range(len(df))
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
            df[price_col] = pd.to_numeric(df[price_col], errors="coerce")
            df = df.dropna(subset=["date", price_col])

            if df.empty:
                raise ValueError(f"{price_col} 全部为空")

            df = (
                df.sort_values(["date", "_raw_order"])
                .drop_duplicates(subset=["date"], keep="last")
                .sort_values("date")
            )

            latest = df.iloc[-1]
            latest_date = latest["date"]
            latest_price = float(latest[price_col])

            return latest_date, latest_price, symbol, price_col

        except Exception as e:
            errors.append(f"{symbol}: {e}")

    raise RuntimeError(f"{item_name} 所有 symbol 都取数失败：{' | '.join(errors)}")


def fetch_latest_futures_daily(
    item_name: str,
    symbols: Iterable[str],
    price_col: str,
) -> tuple[date, float, str, str]:
    """
    从 AKShare 拉取某个期货品种最新日频数据。

    关键逻辑：
    如果 AKShare 同一天返回多条记录：
    - 按原始返回顺序排序；
    - 对 date 去重；
    - keep="last"，保留当天最后一条记录。
    """
    errors = []

    for symbol in dict.fromkeys(item.upper() for item in symbols):
        try:
            df = ak.futures_zh_daily_sina(symbol=symbol)

            if df is None or df.empty:
                raise ValueError("AKShare 返回空表")

            df = normalize_columns(df)

            if "date" not in df.columns:
                raise ValueError(f"缺少 date 列，实际列={list(df.columns)}")

            if price_col not in df.columns:
                raise ValueError(f"缺少 {price_col} 列，实际列={list(df.columns)}")

            df["_raw_order"] = range(len(df))
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
            df = df.dropna(subset=["date"])

            df[price_col] = pd.to_numeric(df[price_col], errors="coerce")

            # 如果使用 settle，但 settle 为空或为 0，尝试回退 close。
            if price_col == "settle" and "close" in df.columns:
                df["close"] = pd.to_numeric(df["close"], errors="coerce")
                df.loc[df["settle"].isna() | (df["settle"] == 0), "settle"] = df["close"]

            df = df.dropna(subset=[price_col])

            if df.empty:
                raise ValueError(f"{price_col} 全部为空")

            # 重点：同一天多条记录，只保留最后一次出现的记录
            df = (
                df.sort_values(["date", "_raw_order"])
                .drop_duplicates(subset=["date"], keep="last")
                .sort_values("date")
            )

            latest = df.iloc[-1]
            latest_date = latest["date"]
            latest_price = float(latest[price_col])

            return latest_date, latest_price, symbol, price_col

        except Exception as e:
            errors.append(f"{symbol}: {e}")

    raise RuntimeError(f"{item_name} 所有 symbol 都取数失败：{' | '.join(errors)}")


def excel_date_key(value) -> str | None:
    """
    将 Excel 单元格日期统一成 YYYY-MM-DD。

    兼容：
    - Excel 序列号；
    - datetime/date；
    - 字符串日期。
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return None

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return pd.to_datetime(value).date().isoformat()
        except Exception:
            return None

    return None


def write_excel_serial_date(cell, trade_date: date):
    """按 Excel 日期序列写入日期。"""
    dt = datetime.combine(trade_date, datetime.min.time())
    cell.value = to_excel(dt)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 2):
    """复制 A/B 列样式，避免插入新行后格式丢失。"""
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)

        if src.has_style:
            dst._style = copy.copy(src._style)

        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)


def find_rows_by_date(ws, date_key: str, start_row: int = 3) -> list[int]:
    """查找某个日期在 sheet 中出现的所有行。"""
    rows = []
    for r in range(start_row, ws.max_row + 1):
        if excel_date_key(ws.cell(r, 1).value) == date_key:
            rows.append(r)
    return rows


def delete_duplicate_dates_keep_top(ws, start_row: int = 3) -> int:
    """
    删除 Excel 表内同一天重复记录。

    因为你的表是：
    - 第 3 行最新；
    - 越往下越旧。

    所以同一个日期如果出现多次，保留最靠上的那条，
    删除下面重复的记录。
    """
    seen = set()
    delete_rows = []

    for r in range(start_row, ws.max_row + 1):
        key = excel_date_key(ws.cell(r, 1).value)

        if not key:
            continue

        if key in seen:
            delete_rows.append(r)
        else:
            seen.add(key)

    for r in reversed(delete_rows):
        ws.delete_rows(r)

    return len(delete_rows)


def upsert_latest_price(ws, trade_date: date, price: float) -> tuple[str, int]:
    """
    写入最新价格。

    情况一：
    第 3 行日期 == AKShare 最新日期
    -> 直接更新 B3，不插入新行。

    情况二：
    AKShare 最新日期 > 第 3 行日期
    -> 插入新的第 3 行，旧数据整体下移。

    情况三：
    AKShare 最新日期 < 第 3 行日期
    -> 不覆盖第 3 行。
    -> 如果这个旧日期在历史行中存在，则更新历史行。
    -> 如果不存在，则跳过。
    """
    new_key = trade_date.isoformat()
    top_key = excel_date_key(ws.cell(3, 1).value)

    if top_key == new_key:
        write_excel_serial_date(ws.cell(3, 1), trade_date)
        ws.cell(3, 2).value = price
        action = "update_B3_same_date"

    else:
        top_date = pd.to_datetime(top_key).date() if top_key else None

        if top_date is None or trade_date > top_date:
            ws.insert_rows(3)

            # 插入后，原第 3 行变成第 4 行。
            # 复制第 4 行样式到第 3 行。
            copy_row_style(ws, src_row=4, dst_row=3, max_col=2)

            write_excel_serial_date(ws.cell(3, 1), trade_date)
            ws.cell(3, 2).value = price
            action = "insert_new_B3"

        else:
            existing_rows = find_rows_by_date(ws, new_key, start_row=3)

            if existing_rows:
                r = existing_rows[0]
                write_excel_serial_date(ws.cell(r, 1), trade_date)
                ws.cell(r, 2).value = price
                action = f"update_existing_row_{r}"
            else:
                action = "skip_older_than_B3"

    deleted = delete_duplicate_dates_keep_top(ws, start_row=3)

    return action, deleted


def fetch_latest_futures_realtime(
    item_name: str,
    symbols: Iterable[str],
    price_col: str = "close",
) -> tuple[date, float, str, str]:
    """Return today's latest minute quote; callers fall back to daily data."""
    errors = []
    today = pd.Timestamp.now(tz="Asia/Shanghai").date()
    for symbol in dict.fromkeys(item.upper() for item in symbols):
        try:
            df = ak.futures_zh_minute_sina(symbol=symbol.upper(), period="1")
            if df is None or df.empty:
                raise ValueError("AKShare 返回空表")
            df = normalize_columns(df)
            latest = df.iloc[-1]
            quote_time = pd.to_datetime(latest["datetime"])
            if quote_time.date() != today:
                raise ValueError(f"最新分钟线仍为 {quote_time.date()}")
            column = price_col if price_col in df.columns else "close"
            return quote_time.date(), float(latest[column]), symbol.upper(), f"realtime_{column}"
        except Exception as exc:
            errors.append(f"{symbol}: {exc}")
    raise ValueError(f"{item_name} 无当日实时行情: {'; '.join(errors)}")


def update_workbook(input_path: Path, output_path: Path) -> list[FetchResult]:
    wb = load_workbook(input_path)
    results: list[FetchResult] = []

    for item_name, cfg in FUTURES_MAP.items():
        sheet_name = resolve_sheet_name(wb, cfg["sheet_candidates"])

        if sheet_name is None:
            logging.warning(
                "[跳过] %s：工作簿中找不到对应 sheet，候选=%s",
                item_name,
                cfg["sheet_candidates"],
            )
            continue

        try:
            try:
                trade_date, price, used_symbol, used_col = fetch_latest_futures_realtime(
                    item_name=item_name,
                    symbols=cfg["symbols"],
                    price_col=cfg["price_col"],
                )
            except Exception as realtime_error:
                logging.info("[%s] 实时行情不可用，回退日线: %s", item_name, realtime_error)
                trade_date, price, used_symbol, used_col = fetch_latest_futures_daily(
                    item_name=item_name,
                    symbols=cfg["symbols"],
                    price_col=cfg["price_col"],
                )

            ws = wb[sheet_name]

            # 更新表头，方便后续知道该 sheet 已经改为 AKShare 口径。
            # 如果你不想改表头，可以注释掉这两行。
            ws.cell(2, 1).value = "指标名称"
            ws.cell(2, 2).value = cfg["desc"]

            action, deleted = upsert_latest_price(ws, trade_date, price)

            results.append(
                FetchResult(
                    item_name=item_name,
                    sheet_name=sheet_name,
                    symbol=used_symbol,
                    trade_date=trade_date,
                    price=price,
                    price_col=used_col,
                    action=action,
                    deleted_duplicates=deleted,
                    desc=cfg["desc"],
                )
            )

            logging.info(
                "[成功] %-12s sheet=%s symbol=%s date=%s price=%s action=%s deleted_dup=%s",
                item_name,
                sheet_name,
                used_symbol,
                trade_date,
                price,
                action,
                deleted,
            )

        except Exception as e:
            logging.exception("[失败] %s: %s", item_name, e)


    for item_name, cfg in FOREIGN_FUTURES_MAP.items():
        try:
            sheet_name, created = resolve_or_create_sheet(
                wb,
                cfg["sheet_candidates"],
                cfg["default_sheet_name"],
            )

            trade_date, price, used_symbol, used_col = fetch_latest_foreign_futures_daily(
                item_name=item_name,
                symbols=cfg["symbols"],
                price_col=cfg["price_col"],
            )

            ws = wb[sheet_name]

            ws.cell(2, 1).value = "指标名称"
            ws.cell(2, 2).value = cfg["desc"]

            action, deleted = upsert_latest_price(ws, trade_date, price)
            if created:
                action = f"create_sheet_and_{action}"

            results.append(
                FetchResult(
                    item_name=item_name,
                    sheet_name=sheet_name,
                    symbol=used_symbol,
                    trade_date=trade_date,
                    price=price,
                    price_col=used_col,
                    action=action,
                    deleted_duplicates=deleted,
                    desc=cfg["desc"],
                )
            )

            logging.info(
                "[成功] %-12s sheet=%s symbol=%s date=%s price=%s action=%s deleted_dup=%s",
                item_name,
                sheet_name,
                used_symbol,
                trade_date,
                price,
                action,
                deleted,
            )

        except Exception as e:
            logging.exception("[失败] %s: %s", item_name, e)

    wb.save(output_path)

    return results


def save_log(results: list[FetchResult], log_path: Path):
    """保存本次更新日志。"""
    rows = [
        {
            "item_name": r.item_name,
            "sheet_name": r.sheet_name,
            "symbol": r.symbol,
            "trade_date": r.trade_date.isoformat(),
            "price": r.price,
            "price_col": r.price_col,
            "action": r.action,
            "deleted_duplicates": r.deleted_duplicates,
            "desc": r.desc,
        }
        for r in results
    ]

    pd.DataFrame(rows).to_csv(log_path, index=False, encoding="utf-8-sig")


def main():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--input",
        required=True,
        help='输入 Excel 文件路径，例如 "price_summarized_optimized.xlsx"',
    )

    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 文件路径。不填则生成 *_akshare.xlsx",
    )

    parser.add_argument(
        "--inplace",
        action="store_true",
        help="直接覆盖原文件。会自动生成 .bak 备份。",
    )

    args = parser.parse_args()

    input_path = Path(args.input).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"找不到输入文件：{input_path}")

    if args.inplace:
        backup_path = input_path.with_suffix(
            input_path.suffix + f".bak_{datetime.now():%Y%m%d_%H%M%S}"
        )
        shutil.copy2(input_path, backup_path)
        output_path = input_path
        print(f"已备份原文件：{backup_path}")
    else:
        output_path = (
            Path(args.output).resolve()
            if args.output
            else input_path.with_name(f"{input_path.stem}_akshare{input_path.suffix}")
        )

    log_path = output_path.with_suffix(".akshare_update_log.csv")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    results = update_workbook(input_path=input_path, output_path=output_path)
    save_log(results, log_path)

    print(f"\n更新完成：{output_path}")
    print(f"日志文件：{log_path}")

    if results:
        print("\n本次成功更新：")
        for r in results:
            print(
                f"- {r.item_name}: "
                f"date={r.trade_date}, "
                f"price={r.price}, "
                f"symbol={r.symbol}, "
                f"sheet={r.sheet_name}, "
                f"action={r.action}, "
                f"删除重复={r.deleted_duplicates}"
            )
    else:
        print("\n没有成功更新任何品种，请查看报错日志。")


if __name__ == "__main__":
    main()
