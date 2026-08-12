from __future__ import annotations

from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


API_URL = "https://www.shpgx.com/marketzhishu/list/3/22"
WORKBOOK_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "液化天然气"
SOURCE_LABEL = "中国LNG出厂价格(全国)-上海石油天然气交易中心"

PAGE_REFERER = "https://www.shpgx.com/html/qgjg.html"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)


def fetch_latest_price():
    response = requests.post(
        API_URL,
        headers={"User-Agent": USER_AGENT, "Referer": PAGE_REFERER},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()

    raw_date = str(payload.get("DATA") or "").strip()
    raw_price = str(payload.get("BASEPRICE") or "").strip()
    if not raw_date or not raw_price:
        raise ValueError("shpgx returned an incomplete payload")

    trade_date = datetime.strptime(raw_date, "%Y-%m-%d")
    price = float(raw_price)
    return trade_date, price


def insert_latest_row(workbook_path, trade_date, price):
    workbook = load_workbook(workbook_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Workbook has no sheet named " + SHEET_NAME)

    worksheet = workbook[SHEET_NAME]

    action = "inserted"
    first_data_row = None

    for row_index in range(3, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_index, column=1).value
        if cell_value is None or cell_value == "":
            continue
        existing_date = cell_value.date() if isinstance(cell_value, datetime) else None
        if existing_date is None:
            continue
        if existing_date == trade_date.date():
            worksheet.cell(row=row_index, column=1).value = trade_date
            worksheet.cell(row=row_index, column=2).value = price
            action = "updated"
            first_data_row = row_index
            break
        if existing_date < trade_date.date() and first_data_row is None:
            first_data_row = row_index

    if first_data_row is None:
        first_data_row = worksheet.max_row + 1

    if action == "inserted":
        worksheet.insert_rows(first_data_row, amount=1)
        new_row = first_data_row
    else:
        new_row = first_data_row

    if action == "updated":
        ref_date_cell = worksheet.cell(row=new_row, column=1)
        ref_price_cell = worksheet.cell(row=new_row, column=2)
    else:
        ref_date_cell = worksheet.cell(row=new_row + 1, column=1)
        ref_price_cell = worksheet.cell(row=new_row + 1, column=2)

    new_date_cell = worksheet.cell(row=new_row, column=1, value=trade_date)
    new_price_cell = worksheet.cell(row=new_row, column=2, value=price)

    new_date_cell.number_format = ref_date_cell.number_format or "yyyy-mm-dd"
    new_price_cell.number_format = ref_price_cell.number_format or "#,##0.00_ "

    new_date_cell.font = Font(name="Times New Roman")
    new_price_cell.font = Font(name="Times New Roman")
    new_date_cell.alignment = Alignment(horizontal="right", vertical="center")
    new_price_cell.alignment = Alignment(horizontal="right", vertical="center")

    from io import BytesIO
    from os import replace

    buffer = BytesIO()
    workbook.save(buffer)
    staging_path = workbook_path.with_name(workbook_path.stem + ".staged.xlsx")
    with staging_path.open("wb") as handle:
        handle.write(buffer.getvalue())
    tmp_path = workbook_path.with_suffix(".tmp.xlsx")
    with tmp_path.open("wb") as handle:
        handle.write(buffer.getvalue())
    try:
        replace(tmp_path, workbook_path)
    except PermissionError:
        # Workbook is locked (e.g. WPS / Excel has it open). Surface the
        # staging copy so the user can swap it in once the lock is released.
        print("WARNING: destination workbook is locked by another process")
        print("Staged updated workbook at: " + str(staging_path))
        return "staged:" + str(staging_path)
    finally:
        tmp_path.unlink(missing_ok=True)
        staging_path.unlink(missing_ok=True)
    return action


def main():
    workbook_path = Path(__file__).resolve().parent / WORKBOOK_FILE
    if not workbook_path.exists():
        raise FileNotFoundError("Workbook not found: " + str(workbook_path))

    trade_date, price = fetch_latest_price()
    action = insert_latest_row(workbook_path, trade_date, price)

    print("Source: " + SOURCE_LABEL)
    print("Endpoint: " + API_URL)
    print("Updated workbook: " + str(workbook_path))
    print("Sheet: " + SHEET_NAME)
    print("Trade date: " + trade_date.date().isoformat() + " (" + action + ")")
    print("LNG price: " + format(price, ",.1f") + " yuan/ton")
    if action.startswith("staged:"):
        print("Next step: close WPS/Excel, then run:")
        print("    python fetch_shpgx_lng.py --commit " + action.split(":", 1)[1])


if __name__ == "__main__":
    main()
