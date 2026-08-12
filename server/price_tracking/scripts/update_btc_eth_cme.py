# -*- coding: utf-8 -*-
from __future__ import annotations

import datetime as dt
import sys
import time
from pathlib import Path
from typing import Iterable

import openpyxl
import requests
from openpyxl.styles import Alignment, Font


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"

# Each entry maps a Yahoo symbol to (sheet_name, layout).
# "ohlcv" writes the full daily bar (date, volume, close, low, high, open).
# "settlement" writes only the close price and preserves the indicator-name
# header labels that sit in row 2 of the ICE Brent sheet.
SHEET_TARGETS = {
    "BTC-USD": ("BTC.CME", "ohlcv"),
    "ETH-USD": ("ETH.CME", "ohlcv"),
    "BZ=F": ("ICE布油", "settlement"),
}

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0 Safari/537.36"
)


def fetch_history(symbol):
    # Return rows as (date, open, high, low, close, volume) sorted descending by
    # date. The final Yahoo bar is intentionally retained as the current quote.
    now = dt.datetime.now()
    period2 = int(time.mktime((now.year, now.month, now.day + 1, 0, 0, 0, 0, 0, 0)))
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    params = {"period1": 0, "period2": period2, "interval": "1d", "events": "history"}
    resp = requests.get(url, params=params, headers={"User-Agent": USER_AGENT}, timeout=120)
    resp.raise_for_status()
    payload = resp.json()
    result = payload["chart"]["result"][0]
    meta = result["meta"]
    print(
        f"  fetched {symbol}: exchange={meta.get('fullExchangeName')}, "
        f"currency={meta.get('currency')}, gmtoffset={meta.get('gmtoffset')}"
    )
    timestamps = result["timestamp"]
    quote = result["indicators"]["quote"][0]

    complete = timestamps

    rows = []
    for ts, o, h, l, c, v in zip(
        complete,
        quote["open"],
        quote["high"],
        quote["low"],
        quote["close"],
        quote["volume"],
    ):
        if o is None or h is None or l is None or c is None:
            continue
        # Use the UTC date as the trading date.
        date = dt.datetime.fromtimestamp(ts, dt.timezone.utc).date()
        rows.append((date, float(o), float(h), float(l), float(c), int(v or 0)))

    rows.sort(key=lambda r: r[0], reverse=True)
    return rows


def _apply_font_alignment(cell):
    cell.font = Font(name="Times New Roman", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_ohlcv(ws, rows):
    # Replace the data block (rows 2..N) with rows for sheets whose header
    # lives in row 1.
    rows = list(rows)
    existing_max_row = ws.max_row
    if existing_max_row >= 2:
        ws.delete_rows(2, existing_max_row - 1)

    number_fmt = "#,##0.00"
    for offset, (date, o, h, l, c, v) in enumerate(rows):
        r = 2 + offset
        ws.cell(row=r, column=1, value=dt.datetime(date.year, date.month, date.day))
        ws.cell(row=r, column=1).number_format = "mm-dd-yy"
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=2).number_format = "General"
        ws.cell(row=r, column=3, value=c)
        ws.cell(row=r, column=3).number_format = number_fmt
        ws.cell(row=r, column=4, value=l)
        ws.cell(row=r, column=4).number_format = number_fmt
        ws.cell(row=r, column=5, value=h)
        ws.cell(row=r, column=5).number_format = number_fmt
        ws.cell(row=r, column=6, value=o)
        ws.cell(row=r, column=6).number_format = number_fmt
        for col in range(1, 7):
            _apply_font_alignment(ws.cell(row=r, column=col))

    return len(rows)


def write_settlement(ws, rows):
    # Replace the data block (rows 3..N) for the ICE Brent sheet, which keeps
    # a 2-column (date, settlement) layout and reserves row 2 for the
    # indicator-name label.
    rows = list(rows)
    existing_max_row = ws.max_row
    if existing_max_row >= 3:
        ws.delete_rows(3, existing_max_row - 2)

    number_fmt = "#,##0.00_ "
    for offset, (date, _o, _h, _l, c, _v) in enumerate(rows):
        r = 3 + offset
        ws.cell(row=r, column=1, value=dt.datetime(date.year, date.month, date.day))
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=c)
        ws.cell(row=r, column=2).number_format = number_fmt
        for col in (1, 2):
            _apply_font_alignment(ws.cell(row=r, column=col))

    return len(rows)


WRITERS = {
    "ohlcv": write_ohlcv,
    "settlement": write_settlement,
}


def main():
    if not WORKBOOK_PATH.exists():
        print(f"workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    for symbol, (sheet_name, layout) in SHEET_TARGETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  sheet {sheet_name!r} missing; skipping", file=sys.stderr)
            continue
        print(f"updating {sheet_name} from {symbol} ({layout})")
        rows = fetch_history(symbol)
        ws = wb[sheet_name]
        written = WRITERS[layout](ws, rows)
        first = rows[0][0] if rows else None
        last = rows[-1][0] if rows else None
        print(f"  wrote {written} rows; range {last} -> {first}")

    wb.save(WORKBOOK_PATH)
    print("saved", WORKBOOK_PATH)
    return 0


if __name__ == "__main__":
    sys.exit(main())
