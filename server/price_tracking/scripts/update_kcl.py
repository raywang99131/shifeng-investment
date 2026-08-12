"""100ppi \u6c2f\u5316\u94be\u57fa\u51c6\u4ef7 -> price_summarized_optimized.xlsx \u6c2f\u5316\u94be\u5de5\u4f5c\u8868\u3002

\u8fd0\u884c\u540e\u4f1a\uff1a
1. \u4ece\u751f\u610f\u793e\uff08100ppi\uff09\u6293\u53d6\u6c2f\u5316\u94be\u57fa\u51c6\u4ef7\u8d70\u52bf\uff1b
2. \u8bfb\u53d6 price_summarized_optimized.xlsx \u4e2d\u201c\u6c2f\u5316\u94be\u201d\u5de5\u4f5c\u8868\u73b0\u6709\u8bb0\u5f55\uff1b
3. \u6309\u65e5\u671f\u5408\u5e76\uff0c\u91cd\u5199\u8be5\u5de5\u4f5c\u8868\u4f7f\u5176\u4e0e\u4ef7\u683c\u603b\u8868 (row 7) \u7684\u201c\u4eca\u65e5\u4ef7\u683c/$B$3\u201d\u3001\u201c\u6628\u65e5\u4ef7\u683c/$B$4\u201d\u8fd0\u7b97\u7ed3\u679c\u4e00\u81f4\uff1b
4. \u540c\u6b65\u4ef7\u683c\u603b\u8868\u4e2d\u201c\u6c2f\u5316\u94be\u201d\u4e00\u884c\u7684\u201c\u6765\u6e90\u201d\u5217\u4e3a\u201c100ppi(\u751f\u610f\u793e)\u201d\u3002
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.datetime import from_excel


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "\u6c2f\u5316\u94be"
HEADER_LABEL = "\u6307\u6807\u540d\u79f0"
INDICATOR_LABEL = "中国:华东地区:市场价(平均价):氯化钾粉(62%)"
SUMMARY_SOURCE_LABEL = "100ppi(\u751f\u610f\u793e)"

DEFAULT_LOOKBACK_DAYS = 30
HISTORY_LOOKBACK_DAYS = 365
REQUEST_TIMEOUT = 20


@dataclass(frozen=True)
class SheetConfig:
    sheet_name: str
    indicator_label: str
    summary_source_label: str


CONFIG = SheetConfig(
    sheet_name=SHEET_NAME,
    indicator_label=INDICATOR_LABEL,
    summary_source_label=SUMMARY_SOURCE_LABEL,
)


# -----------------------------------------------------------------------
# 100ppi \u6293\u53d6
# -----------------------------------------------------------------------

def parse_mmdd_to_date(mmdd: str):
    """\u628a 07-12 \u8f6c\u4e3a\u5f53\u524d\u5e74\u4efd\u65e5\u671f\uff1b\u5982\u679c\u89e3\u6790\u51fa\u6765\u662f\u672a\u6765\u65e5\u671f\uff0c\u5219\u81ea\u52a8\u56de\u9000\u4e00\u5e74\u3002"""
    mmdd = str(mmdd).strip()
    m = re.match(r"^(\d{1,2})-(\d{1,2})$", mmdd)
    if not m:
        return pd.NaT

    year = datetime.now().year
    month = int(m.group(1))
    day = int(m.group(2))

    dt = pd.Timestamp(year=year, month=month, day=day)
    today = pd.Timestamp.today().normalize()

    if dt > today + pd.Timedelta(days=3):
        dt = pd.Timestamp(year=year - 1, month=month, day=day)

    return dt


def fetch_kcl_from_100ppi() -> pd.DataFrame:
    """\u4ece\u751f\u610f\u793e\u6293\u53d6\u6c2f\u5316\u94be\u57fa\u51c6\u4ef7\u8d70\u52bf\uff0c\u5355\u4f4d\uff1a\u5143/\u5428"""
    commodity = quote("\u6c2f\u5316\u94be")
    url = f"https://m1.100ppi.com/vane/759-{commodity}.html"

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
    }

    resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding

    soup = BeautifulSoup(resp.text, "lxml")
    text = soup.get_text("\n", strip=True)

    pattern = re.compile(
        r"(?P<date>\d{1,2}-\d{1,2})\s+"
        r"(?P<price>\d+(?:\.\d+)?)\s+"
        r"(?P<change>[+-]?\d+(?:\.\d+)?)%"
    )

    rows = []
    for m in pattern.finditer(text):
        rows.append({
            "date": parse_mmdd_to_date(m.group("date")),
            "price": float(m.group("price")),
        })

    if not rows:
        with open("debug_100ppi_kcl.html", "w", encoding="utf-8") as f:
            f.write(resp.text)
        raise ValueError(
            "\u672a\u4ece\u751f\u610f\u793e\u9875\u9762\u89e3\u6790\u5230\u6c2f\u5316\u94be\u4ef7\u683c\u3002"
            "\u5df2\u4fdd\u5b58 debug_100ppi_kcl.html\uff0c\u8bf7\u68c0\u67e5\u662f\u5426\u88ab\u53cd\u722c\u6216\u9875\u9762\u7ed3\u6784\u53d8\u5316\u3002"
        )

    df = pd.DataFrame(rows)
    df = df.dropna(subset=["date", "price"])
    df = df.drop_duplicates(subset=["date"], keep="first")
    df = df.sort_values("date").reset_index(drop=True)
    return df


def get_latest_kcl_price() -> dict[str, Any]:
    """\u900f\u51fa\u6700\u8fd1\u4e00\u5929\u53ca\u524d\u4e00\u5929\u7684\u4ef7\u683c\u6982\u89c8\uff0c\u4ec5\u7528\u4e8e\u8c03\u8bd5/CLI \u67e5\u770b\u3002"""
    df = fetch_kcl_from_100ppi().dropna(subset=["price"]).sort_values("date")

    latest = df.iloc[-1]
    prev = df.iloc[-2] if len(df) >= 2 else None

    result = {
        "date": latest["date"].date(),
        "today_price": float(latest["price"]),
        "unit": "\u5143/\u5428",
        "source": "\u751f\u610f\u793e-\u6c2f\u5316\u94be\u57fa\u51c6\u4ef7",
    }
    if prev is not None:
        result["yesterday_price"] = float(prev["price"])
        result["change"] = float(latest["price"] - prev["price"])
        result["pct_change"] = float((latest["price"] - prev["price"]) / prev["price"])
    return result


# -----------------------------------------------------------------------
# \u5de5\u4f5c\u8868\u5408\u5e76 / \u5199\u5165
# -----------------------------------------------------------------------

def normalize_trade_date(value: Any) -> date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        try:
            return pd.to_datetime(text).date()
        except Exception:
            return None
    return None


def normalize_price(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_existing_rows(worksheet) -> list[list[Any]]:
    rows: list[list[Any]] = []
    if worksheet is None:
        return rows
    for row in worksheet.iter_rows(min_row=3, max_col=2, values_only=True):
        trade_date = normalize_trade_date(row[0])
        price = normalize_price(row[1])
        if trade_date is None or price is None:
            continue
        rows.append([trade_date, price])
    return rows


def detect_fetch_start_date(worksheet) -> date:
    today = date.today()
    existing_rows = read_existing_rows(worksheet)
    if not existing_rows:
        return today - timedelta(days=HISTORY_LOOKBACK_DAYS)
    latest_existing = max(row[0] for row in existing_rows)
    return latest_existing - timedelta(days=DEFAULT_LOOKBACK_DAYS)


def fetch_kcl_rows(start_date: date, end_date: date) -> list[list[Any]]:
    df = fetch_kcl_from_100ppi()
    if df.empty:
        return []
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    df = df[(df["date"] >= start_ts) & (df["date"] <= end_ts)]
    out = []
    for _, row in df.iterrows():
        ts = row["date"]
        out.append([ts.date() if hasattr(ts, "date") else ts, float(row["price"])])
    return out


def merge_rows(existing_rows, new_rows):
    rows_by_date = {row[0]: row for row in existing_rows}
    inserted = 0
    updated = 0
    for row in new_rows:
        if row[0] in rows_by_date:
            updated += 1
        else:
            inserted += 1
        rows_by_date[row[0]] = row
    merged_rows = sorted(rows_by_date.values(), key=lambda item: item[0], reverse=True)
    return merged_rows, inserted, updated


def ensure_sheet(workbook):
    if CONFIG.sheet_name in workbook.sheetnames:
        return workbook[CONFIG.sheet_name]
    return workbook.create_sheet(CONFIG.sheet_name)


def rewrite_sheet(worksheet, rows: list[list[Any]]) -> None:
    if worksheet.max_row > 0:
        worksheet.delete_rows(1, worksheet.max_row)

    worksheet["A1"] = None
    worksheet["B1"] = None
    worksheet["A2"] = HEADER_LABEL
    worksheet["B2"] = CONFIG.indicator_label

    times_font = Font(name="Times New Roman")
    worksheet["A2"].font = times_font
    worksheet["B2"].font = times_font

    for trade_date, price in rows:
        worksheet.append([trade_date, price])

    for row in worksheet.iter_rows(min_row=3, max_col=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = times_font

    for cell in worksheet["A"][2:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in worksheet["B"][2:]:
        cell.number_format = "0.00"

    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 38


def save_workbook(workbook, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError:
        temp_path = workbook_path.with_name(f"{workbook_path.stem}.tmp{workbook_path.suffix}")
        workbook.save(temp_path)
        temp_path.replace(workbook_path)


def update_summary_source(workbook, source_label: str) -> bool:
    """\u540c\u6b65\u4ef7\u683c\u603b\u8868\u4e2d\u201c\u6c2f\u5316\u94be\u201d\u4e00\u884c\u7684\u201c\u6765\u6e90\u201d\u5217\uff0c\u4f7f\u5176\u4e0e\u5b9e\u9645\u62c9\u53d6\u6e90\u4e00\u81f4\u3002"""
    if "\u603b\u8868" not in workbook.sheetnames:
        return False
    sheet = workbook["\u603b\u8868"]
    for row in sheet.iter_rows(min_row=2, values_only=False):
        name_cell = row[0]
        if name_cell.value is None:
            continue
        if str(name_cell.value).strip() == CONFIG.sheet_name:
            if len(row) >= 3:
                row[2].value = source_label
            return True
    return False


# -----------------------------------------------------------------------
# \u4e3b\u6d41\u7a0b
# -----------------------------------------------------------------------

def update_kcl_workbook(
    workbook_path: Path = WORKBOOK_PATH,
    end_date: date | None = None,
) -> dict[str, Any]:
    if end_date is None:
        end_date = date.today()

    workbook = load_workbook(workbook_path)
    existing_sheet = (
        workbook[CONFIG.sheet_name]
        if CONFIG.sheet_name in workbook.sheetnames
        else None
    )
    start_date = detect_fetch_start_date(existing_sheet)
    new_rows = fetch_kcl_rows(start_date=start_date, end_date=end_date)
    existing_rows = read_existing_rows(existing_sheet)
    merged_rows, inserted, updated = merge_rows(existing_rows, new_rows)

    worksheet = ensure_sheet(workbook)
    rewrite_sheet(worksheet, merged_rows)

    source_updated = update_summary_source(workbook, CONFIG.summary_source_label)

    save_workbook(workbook, workbook_path)

    return {
        "workbook_path": str(workbook_path.resolve()),
        "total_rows": len(merged_rows),
        "inserted": inserted,
        "updated": updated,
        "fetch_start": start_date.isoformat(),
        "fetch_end": end_date.isoformat(),
        "source_summary_updated": source_updated,
    }


def main() -> None:
    summary = update_kcl_workbook()
    print(f"\u66f4\u65b0\u5de5\u4f5c\u7c3f\uff1a{summary['workbook_path']}")
    print(
        f"{CONFIG.sheet_name}: total={summary['total_rows']}, "
        f"inserted={summary['inserted']}, updated={summary['updated']}, "
        f"fetch_window={summary['fetch_start']}~{summary['fetch_end']}, "
        f"summary_source_updated={summary['source_summary_updated']}"
    )


if __name__ == "__main__":
    main()
