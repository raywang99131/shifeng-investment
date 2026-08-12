#!/usr/bin/env python3
"""Backfill official CTIA tungsten daily prices into the tracking workbook."""

from __future__ import annotations

import html
import os
import re
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook


WORKBOOK = Path(__file__).resolve().parents[1] / "price_summarized_optimized.xlsx"
API_URL = "https://www.ctia.com.cn/wp-json/wp/v2/posts"
SHEETS = {"black": "黑钨精矿", "waste": "废钨棒材", "powder": "钨粉"}


def clean_html(value: str) -> str:
    value = html.unescape(re.sub(r"<[^>]+>", " ", value or ""))
    return re.sub(r"\s+", " ", value).strip()


def parse_snapshot(post: dict) -> dict | None:
    text = clean_html(post.get("content", {}).get("rendered", ""))
    if "钨市场行情" not in text:
        return None
    patterns = {
        "black": r"65%黑钨精矿价格\s*([0-9,.]+)\s*万元/标吨",
        "waste": r"废钨棒材价格\s*([0-9,.]+)\s*元/千克",
        "powder": r"(?:^|[\s，。])钨粉价格\s*([0-9,.]+)\s*元/千克",
    }
    values = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if not match:
            return None
        values[key] = float(match.group(1).replace(",", ""))
    return {
        "date": datetime.fromisoformat(post["date"]).date(),
        "values": values,
    }


def fetch_snapshots(days: int = 60) -> list[dict]:
    start = date.today() - timedelta(days=days)
    params = {
        "categories": 11,
        "after": datetime.combine(start, time.min, tzinfo=timezone.utc).isoformat(),
        "before": datetime.combine(date.today() + timedelta(days=1), time.min, tzinfo=timezone.utc).isoformat(),
        "per_page": 100,
        "page": 1,
        "_fields": "date,link,content",
    }
    snapshots = {}
    while True:
        response = requests.get(API_URL, params=params, timeout=30)
        response.raise_for_status()
        for post in response.json():
            snapshot = parse_snapshot(post)
            if snapshot:
                snapshots[snapshot["date"]] = snapshot
        if params["page"] >= int(response.headers.get("X-WP-TotalPages", "1")):
            break
        params["page"] += 1
    return [snapshots[key] for key in sorted(snapshots, reverse=True)]


def replace_sheet_history(ws, values: dict[date, float]) -> None:
    existing = {}
    for raw_date, raw_value in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        if raw_date is None or raw_value is None:
            continue
        parsed_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        existing[parsed_date] = float(raw_value)
    existing.update(values)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    for day in sorted(existing, reverse=True):
        ws.append([datetime.combine(day, time.min), existing[day]])


def main() -> None:
    snapshots = fetch_snapshots()
    if not snapshots:
        raise RuntimeError("中钨在线近 60 天未解析到钨价历史")
    workbook = load_workbook(WORKBOOK)
    for key, sheet_name in SHEETS.items():
        replace_sheet_history(
            workbook[sheet_name],
            {snapshot["date"]: snapshot["values"][key] for snapshot in snapshots},
        )
    temporary = WORKBOOK.with_suffix(".backfill.tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, WORKBOOK)
    print(
        f"中钨在线历史回填完成: {len(snapshots)} 个交易日, "
        f"{snapshots[-1]['date']} 至 {snapshots[0]['date']}"
    )


if __name__ == "__main__":
    main()
