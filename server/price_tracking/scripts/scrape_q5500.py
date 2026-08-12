from __future__ import annotations

from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


API_URL = "https://www.cctd.com.cn/Echarts/data/HBHCKJ.php"
WORKBOOK_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "Q5500动力煤"
HEADERS = ["时间", "Q5500K", "环比增减", "Q5000K", "环比增减", "Q4500K", "环比增减"]


def fetch_market_data() -> list[dict[str, Any]]:
    response = requests.post(
        API_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            ),
            "Referer": "https://www.cctd.com.cn/",
        },
        timeout=30,
    )
    response.raise_for_status()

    data = response.json()
    if not isinstance(data, list):
        raise ValueError("Unexpected API response; expected a list of rows.")
    return data


def to_int_or_none(value: Any) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def build_rows(raw_rows: list[dict[str, Any]]) -> list[list[Any]]:
    processed_rows: list[list[Any]] = []
    previous_q5500: int | None = None
    previous_q5000: int | None = None
    previous_q4500: int | None = None

    for row in raw_rows:
        trade_date = row.get("name", "")
        q5500 = to_int_or_none(row.get("age"))
        q5000 = to_int_or_none(row.get("product"))
        q4500 = to_int_or_none(row.get("product1"))

        q5500_change = q5500 - previous_q5500 if q5500 is not None and previous_q5500 is not None else ""
        q5000_change = q5000 - previous_q5000 if q5000 is not None and previous_q5000 is not None else ""
        q4500_change = q4500 - previous_q4500 if q4500 is not None and previous_q4500 is not None else ""

        processed_rows.append(
            [
                trade_date,
                q5500 if q5500 is not None else row.get("age", ""),
                q5500_change,
                q5000 if q5000 is not None else row.get("product", ""),
                q5000_change,
                q4500 if q4500 is not None else row.get("product1", ""),
                q4500_change,
            ]
        )

        previous_q5500 = q5500 if q5500 is not None else previous_q5500
        previous_q5000 = q5000 if q5000 is not None else previous_q5000
        previous_q4500 = q4500 if q4500 is not None else previous_q4500

    processed_rows.reverse()
    return processed_rows


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="0B67D4")
    # 标题字体：Times New Roman，白色，加粗
    header_font = Font(name="Times New Roman", color="FFFFFF", bold=True)
    # 正文字体：Times New Roman，默认黑色
    normal_font = Font(name="Times New Roman")
    alt_fill = PatternFill("solid", fgColor="EEF3F6")
    center = Alignment(horizontal="center", vertical="center")

    # 第一行样式
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 数据行样式
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_index % 2 == 0 else None
        for cell in row:
            cell.alignment = center
            cell.font = normal_font
            if fill is not None:
                cell.fill = fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 12
    worksheet.column_dimensions["F"].width = 12
    worksheet.column_dimensions["G"].width = 12


def write_workbook(rows: list[list[Any]], workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]

    worksheet = workbook.create_sheet(title=SHEET_NAME)
    worksheet.append(HEADERS)

    for row in rows:
        worksheet.append(row)

    style_sheet(worksheet)
    workbook.save(workbook_path)


def main() -> None:
    current_dir = Path(__file__).resolve().parent
    workbook_path = current_dir / WORKBOOK_FILE
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    raw_rows = fetch_market_data()
    rows = build_rows(raw_rows)
    write_workbook(rows, workbook_path)

    print(f"Updated workbook: {workbook_path}")
    print(f"Inserted sheet: {SHEET_NAME}")
    print(f"Rows written: {len(rows)}")
    if rows:
        print(f"Latest row: {rows[0][0]} | Q5500K={rows[0][1]}")


if __name__ == "__main__":
    main()
