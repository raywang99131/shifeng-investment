# -*- coding: utf-8 -*-
"""Update the 伦敦金现 sheet from huilvbiao.com XAU page."""
from __future__ import annotations

import json
import sys
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font  # 新增导入


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "伦敦金现"
HEADER_LABEL = "指标名称"
INDICATOR_LABEL = "伦敦金现:IDC"

KLINE_URL = "https://www.huilvbiao.com/api/gold_autd_kline?t=xau"
REAL_URL = "https://www.huilvbiao.com/api/gold_autd_real?t=xau"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0 Safari/537.36"
)


def _request(url: str) -> Any:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Referer": "https://www.huilvbiao.com/gold/xau",
            "Accept": "application/json, text/plain, */*",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())


def fetch_kline() -> list[dict[str, Any]]:
    return _request(KLINE_URL)


def fetch_latest_spot() -> tuple[date, float]:
    bars = _request(REAL_URL)
    if not bars:
        raise RuntimeError("real-time API returned no bars")
    latest = bars[0]
    dt = datetime.strptime(latest["date_time"].split(" ")[0], "%Y-%m-%d").date()
    return dt, float(latest["new"])


def to_rows(payload: list[dict[str, Any]]) -> list[tuple[date, float]]:
    rows: list[tuple[date, float]] = []
    for item in payload:
        day_str = item.get("day")
        if not day_str:
            raise ValueError(f"missing 'day' in payload item: {item}")
        d = datetime.strptime(day_str, "%Y/%m/%d").date()
        rows.append((d, float(item["close"])))
    rows.sort(key=lambda r: r[0], reverse=True)
    return rows


def rewrite_sheet(ws, rows: list[tuple[date, float]]) -> int:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws["A1"] = None
    ws["B1"] = None
    ws["A2"] = HEADER_LABEL
    ws["B2"] = INDICATOR_LABEL

    # 创建 Times New Roman 字体对象
    times_font = Font(name='Times New Roman')
    # 设置标题字体
    ws["A2"].font = times_font
    ws["B2"].font = times_font

    for offset, (d, price) in enumerate(rows):
        r = 3 + offset
        date_cell = ws.cell(row=r, column=1, value=datetime(d.year, d.month, d.day))
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.font = times_font  # 设置日期单元格字体

        price_cell = ws.cell(row=r, column=2, value=price)
        price_cell.number_format = "0.00"
        price_cell.font = times_font  # 设置价格单元格字体

    return len(rows)


def save_workbook(wb, path: Path) -> None:
    try:
        wb.save(path)
    except PermissionError:
        tmp = path.with_name(f"{path.stem}.tmp{path.suffix}")
        wb.save(tmp)
        tmp.replace(path)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    payload = fetch_kline()
    rows = to_rows(payload)
    if not rows:
        print("kline API returned no rows", file=sys.stderr)
        return 1

    latest_date, kline_close = rows[0]
    spot_date, spot_price = fetch_latest_spot()
    if spot_date == latest_date and spot_price != kline_close:
        rows[0] = (latest_date, spot_price)
        print(
            f"Overrode {latest_date} kline close {kline_close} "
            f"with live spot {spot_price}"
        )

    print(
        f"Fetched {len(rows)} kline rows; latest = {rows[0][0]} close={rows[0][1]}"
    )

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"sheet {SHEET_NAME!r} not in workbook", file=sys.stderr)
        return 1
    ws = wb[SHEET_NAME]
    written = rewrite_sheet(ws, rows)
    save_workbook(wb, WORKBOOK_PATH)

    head = [rows[0], rows[1], rows[2]]
    print(f"Updated {SHEET_NAME}: wrote {written} rows; head={head}")
    return 0


if __name__ == "__main__":
    sys.exit(main())