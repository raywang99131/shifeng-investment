from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright


PAGE_URL = "https://zh.tradingeconomics.com/commodity/brent-crude-oil"
WORKBOOK_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SHEET_NAME = "\u539f\u6cb9"
SUMMARY_SHEET = "\u603b\u8868"
SUMMARY_ROW = 32  # \u5e03\u4f26\u7279\u539f\u6cb9\u884c
INDICATOR_NAME = "\u5e03\u4f26\u7279\u539f\u6cb9\uff08TradingEconomics \u65e5\u9891 CFD \u6536\u76d8\u4ef7\uff09"
SOURCE_LABEL = "TradingEconomics"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)

HEADER_FILL = PatternFill("solid", fgColor="0B67D4")
HEADER_FONT = Font(name="Times New Roman", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Times New Roman")
CENTER = Alignment(horizontal="center", vertical="center")


def fetch_chart_series() -> dict[str, Any]:
    """Render the TradingEconomics page and pull the main chart series payload."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=USER_AGENT, locale="zh-CN")
        page = ctx.new_page()
        page.goto(PAGE_URL, wait_until="domcontentloaded", timeout=60000)
        try:
            page.wait_for_selector(".highcharts-container, .highstock", timeout=20000)
        except Exception as exc:  # pragma: no cover - log and continue
            print(f"[warn] chart container wait timed out: {exc!r}")
        page.wait_for_timeout(8000)

        payload = page.evaluate(
            """() => {
            const chart = (window.Highcharts && Highcharts.charts)
                ? Highcharts.charts.find(c => c && c.series && c.series.length)
                : null;
            if (!chart) return null;
            const series = chart.series[0];
            return {
                name: series.name,
                data: series.options.data || [],
            };
        }"""
        )
        browser.close()
    if payload is None:
        raise RuntimeError("Failed to locate Highcharts series on the Brent page.")
    return payload


def build_daily_rows(series_data: list[Any]) -> list[tuple[date, float]]:
    """Filter the chart payload to historical daily closes, ordered DESC by date."""
    rows: list[tuple[date, float]] = []
    for point in series_data:
        if not isinstance(point, dict):
            # \u672b\u5c3e\u7684\u524d\u5411 OHLC \u6570\u7ec4\uff0c\u8df3\u8fc7\u4ee5\u907f\u514d\u6c61\u67d3\u5386\u53f2\u6570\u636e
            continue
        date_str = (point.get("date") or "").strip()
        price = point.get("y")
        if not date_str or price is None:
            continue
        trade_date = datetime.fromisoformat(date_str).date()
        rows.append((trade_date, float(price)))

    rows.sort(key=lambda item: item[0], reverse=True)
    return rows


def recreate_sheet(workbook) -> Any:
    sheet_index = workbook.sheetnames.index(SHEET_NAME) if SHEET_NAME in workbook.sheetnames else None
    if sheet_index is not None:
        del workbook[SHEET_NAME]
    return workbook.create_sheet(title=SHEET_NAME, index=sheet_index)


def write_sheet(worksheet, rows: list[tuple[date, float]]) -> None:
    # \u4fdd\u7559\u539f\u6709\u7ed3\u6784\uff1a\u7b2c 1 \u884c\u7a7a\u767d\uff0c\u7b2c 2 \u884c\u6307\u6807\u540d\u79f0
    worksheet.cell(row=1, column=1, value=None)
    worksheet.cell(row=2, column=1, value="\u6307\u6807\u540d\u79f0")
    worksheet.cell(row=2, column=2, value=INDICATOR_NAME)

    for row_index, (trade_date, price) in enumerate(rows, start=3):
        worksheet.cell(row=row_index, column=1, value=datetime(trade_date.year, trade_date.month, trade_date.day))
        worksheet.cell(row=row_index, column=2, value=price)

    for cell in worksheet[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row_index in range(3, worksheet.max_row + 1):
        for cell in worksheet[row_index]:
            cell.font = BODY_FONT
            cell.alignment = CENTER
        worksheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_index, column=2).number_format = "0.00"

    worksheet.freeze_panes = "A3"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 14


def update_summary_source(workbook) -> None:
    if SUMMARY_SHEET not in workbook.sheetnames:
        return
    summary = workbook[SUMMARY_SHEET]
    summary.cell(row=SUMMARY_ROW, column=3, value=SOURCE_LABEL)


def save_workbook(workbook, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError:
        buffer = BytesIO()
        workbook.save(buffer)
        with workbook_path.open("rb+") as handle:
            handle.seek(0)
            handle.write(buffer.getvalue())
            handle.truncate()


def main() -> None:
    workbook_path = Path(__file__).resolve().parent / WORKBOOK_FILE
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    payload = fetch_chart_series()
    rows = build_daily_rows(payload.get("data") or [])
    if not rows:
        raise RuntimeError("No historical daily closes were parsed from the chart payload.")

    workbook = load_workbook(workbook_path)
    worksheet = recreate_sheet(workbook)
    write_sheet(worksheet, rows)
    update_summary_source(workbook)
    save_workbook(workbook, workbook_path)

    print(f"\u5df2\u66f4\u65b0\u5de5\u4f5c\u7c3f: {workbook_path}")
    print(f"\u66f4\u65b0\u8868\u540d: {SHEET_NAME}\uff08\u603b\u8868\u7b2c {SUMMARY_ROW} \u884c\u6e90\u5df2\u540c\u6b65\u4e3a {SOURCE_LABEL}\uff09")
    print(f"\u5199\u5165\u884c\u6570: {len(rows)}\uff08\u542b\u8868\u5934\uff09")
    latest_date, latest_price = rows[0]
    earliest_date, earliest_price = rows[-1]
    print(f"\u6700\u65b0\u4e00\u884c: {latest_date.isoformat()} = {latest_price:.2f}")
    print(f"\u6700\u65e9\u4e00\u884c: {earliest_date.isoformat()} = {earliest_price:.2f}")


if __name__ == "__main__":
    main()
