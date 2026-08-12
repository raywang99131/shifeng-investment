from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import StringIO
from pathlib import Path
import subprocess
import time
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font          # 新增导入
from openpyxl.utils.datetime import from_excel


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
SEARCH_URL = "https://market.cnal.com/historical/search.html"
REFERER_URL = "https://market.cnal.com/historical/lmeofficial.html"
DEFAULT_LOOKBACK_DAYS = 30
HISTORY_LOOKBACK_DAYS = 365

COL_PRODUCT = "\u54c1\u540d"
COL_SETTLEMENT = "\u7ed3\u7b97\u4ef7"
COL_DATE = "\u65e5\u671f"
HEADER_LABEL = "\u6307\u6807\u540d\u79f0"


@dataclass(frozen=True)
class SheetConfig:
    sheet_name: str
    product_name: str
    select_id: str
    indicator_label: str


SHEETS = [
    SheetConfig(
        sheet_name="LME\u94dc",
        product_name="LME\u94dc",
        select_id="131",
        indicator_label="\u73b0\u8d27\u7ed3\u7b97\u4ef7:LME\u94dc",
    ),
    SheetConfig(
        sheet_name="LME\u94dd",
        product_name="LME\u539f\u94dd",
        select_id="134",
        indicator_label="\u73b0\u8d27\u7ed3\u7b97\u4ef7:LME\u94dd",
    ),
]


def load_or_create_workbook(workbook_path: Path):
    if workbook_path.exists():
        return load_workbook(workbook_path)

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    return workbook


def normalize_trade_date(value: Any) -> date | None:
    if value in {None, ""}:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass

        try:
            return pd.to_datetime(text).date()
        except Exception:
            return None

    return None


def read_existing_rows(worksheet) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=3, max_col=2, values_only=True):
        trade_date = normalize_trade_date(row[0])
        if trade_date is None:
            continue
        rows.append([trade_date, row[1]])
    return rows


def detect_fetch_start_date(worksheet: Any | None) -> date:
    if worksheet is None:
        return date.today() - timedelta(days=HISTORY_LOOKBACK_DAYS)

    existing_rows = read_existing_rows(worksheet)
    if not existing_rows:
        return date.today() - timedelta(days=HISTORY_LOOKBACK_DAYS)

    latest_existing = max(row[0] for row in existing_rows)
    return latest_existing - timedelta(days=DEFAULT_LOOKBACK_DAYS)


def fetch_cnal_history(config: SheetConfig, start_date: date, end_date: date) -> list[list[Any]]:
    payload = {
        "starttime": start_date.isoformat(),
        "endtime": end_date.isoformat(),
        "selectid": config.select_id,
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": REFERER_URL,
    }

    request_error: Exception | None = None
    html = ""
    for attempt in range(2):
        try:
            response = requests.post(
                SEARCH_URL,
                data=payload,
                headers=headers,
                timeout=8,
            )
            response.raise_for_status()
            html = response.text
            break
        except requests.RequestException as exc:
            request_error = exc
            if attempt < 1:
                time.sleep(attempt + 1)

    if not html:
        command = [
            "curl",
            "--fail",
            "--silent",
            "--show-error",
            "--http1.1",
            "--retry",
            "1",
            "--connect-timeout",
            "5",
            "--max-time",
            "12",
            "--user-agent",
            headers["User-Agent"],
            "--referer",
            headers["Referer"],
        ]
        for key, value in payload.items():
            command.extend(["--data-urlencode", f"{key}={value}"])
        command.append(SEARCH_URL)

        try:
            completed = subprocess.run(
                command,
                check=True,
                capture_output=True,
                text=True,
                timeout=15,
            )
            html = completed.stdout
        except (subprocess.SubprocessError, OSError) as curl_error:
            raise RuntimeError(
                f"CNAL requests and curl both failed: requests={request_error}; curl={curl_error}"
            ) from curl_error

    tables = pd.read_html(StringIO(html))
    if not tables:
        raise RuntimeError(f"{config.sheet_name} did not return any tables")

    df = tables[0].copy()
    expected_cols = {COL_PRODUCT, COL_SETTLEMENT, COL_DATE}
    if not expected_cols.issubset(set(df.columns)):
        raise RuntimeError(f"{config.sheet_name} returned unexpected columns: {list(df.columns)}")

    df = df[[COL_PRODUCT, COL_SETTLEMENT, COL_DATE]].copy()
    df = df[df[COL_PRODUCT].astype(str).str.strip() == config.product_name]
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], format="%y-%m-%d", errors="coerce").dt.date
    df[COL_SETTLEMENT] = pd.to_numeric(df[COL_SETTLEMENT], errors="coerce")
    df = df.dropna(subset=[COL_DATE, COL_SETTLEMENT])

    return [[row[COL_DATE], float(row[COL_SETTLEMENT])] for _, row in df.iterrows()]


def merge_rows(existing_rows: list[list[Any]], new_rows: list[list[Any]]) -> tuple[list[list[Any]], int, int]:
    rows_by_date: dict[date, list[Any]] = {row[0]: row for row in existing_rows}
    inserted = 0
    updated = 0

    for row in new_rows:
        if row[0] in rows_by_date:
            updated += 1
        else:
            inserted += 1
        rows_by_date[row[0]] = row

    merged_rows = sorted(rows_by_date.values(), key=lambda item: item[0], reverse=True)
    return merged_rows, inserted, updated


def ensure_sheet(workbook, config: SheetConfig):
    if config.sheet_name in workbook.sheetnames:
        return workbook[config.sheet_name]
    return workbook.create_sheet(config.sheet_name)


def rewrite_sheet(worksheet, config: SheetConfig, rows: list[list[Any]]) -> None:
    if worksheet.max_row > 0:
        worksheet.delete_rows(1, worksheet.max_row)

    worksheet["A1"] = None
    worksheet["B1"] = None
    worksheet["A2"] = HEADER_LABEL
    worksheet["B2"] = config.indicator_label

    # ---------- 字体设置开始 ----------
    times_font = Font(name='Times New Roman')
    worksheet["A2"].font = times_font
    worksheet["B2"].font = times_font
    # ---------- 字体设置结束 ----------

    for trade_date, price in rows:
        worksheet.append([trade_date, price])

    # ---------- 为所有数据单元格设置字体 ----------
    for row in worksheet.iter_rows(min_row=3, max_col=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = times_font
    # ---------- 字体设置结束 ----------

    for cell in worksheet["A"][2:]:
        cell.number_format = "yyyy-mm-dd"

    for cell in worksheet["B"][2:]:
        cell.number_format = "0.00"


def save_workbook(workbook, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError:
        temp_path = workbook_path.with_name(f"{workbook_path.stem}.tmp{workbook_path.suffix}")
        workbook.save(temp_path)
        temp_path.replace(workbook_path)


def update_sheet(workbook, config: SheetConfig, end_date: date) -> tuple[int, int, int]:
    existing_sheet = workbook[config.sheet_name] if config.sheet_name in workbook.sheetnames else None
    start_date = detect_fetch_start_date(existing_sheet)
    new_rows = fetch_cnal_history(config, start_date=start_date, end_date=end_date)
    existing_rows = read_existing_rows(existing_sheet) if existing_sheet is not None else []
    merged_rows, inserted, updated = merge_rows(existing_rows, new_rows)

    worksheet = ensure_sheet(workbook, config)
    rewrite_sheet(worksheet, config, merged_rows)

    return len(merged_rows), inserted, updated


def main() -> None:
    workbook = load_or_create_workbook(WORKBOOK_PATH)
    end_date = date.today()

    summaries: list[str] = []
    for config in SHEETS:
        total_rows, inserted, updated = update_sheet(workbook, config, end_date=end_date)
        summaries.append(
            f"{config.sheet_name}: total={total_rows}, inserted={inserted}, updated={updated}"
        )

    save_workbook(workbook, WORKBOOK_PATH)

    print(f"Updated workbook: {WORKBOOK_PATH.resolve()}")
    for summary in summaries:
        print(summary)


if __name__ == "__main__":
    main()
