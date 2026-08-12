from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


API_URL = "https://www.sci99.com/priceMonitor/listProductPagePrice"
TARGET_FILE = Path(__file__).resolve().parent.parent / "price_summarized_optimized.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="2F57D4")
HEADER_FONT = Font(name="Times New Roman", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Times New Roman")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


@dataclass(frozen=True)
class ProductConfig:
    name: str
    page_url: str
    old_id: str
    sheet_name: str
    preamble_rows: list[list[Any]]
    column_widths: dict[str, float]


PRODUCTS = [
    ProductConfig(
        name="R32",
        page_url="https://www.sci99.com/monitor-1572-0.html",
        old_id="1572",
        sheet_name="制冷剂R32",
        preamble_rows=[["日期", "价格"]],
        column_widths={"A": 16, "B": 14},
    ),
    ProductConfig(
        name="TDI",
        page_url="https://www.sci99.com/monitor-375-0.html",
        old_id="375",
        sheet_name="TDI",
        preamble_rows=[[None, None], ["指标名称", "中国:现货价:甲苯二异氰酸酯"]],
        column_widths={"A": 16, "B": 18},
    ),
    ProductConfig(
        name="Pure MDI",
        page_url="https://www.sci99.com/monitor-94717214-1.html",
        old_id="94717214",
        sheet_name="MDI",
        preamble_rows=[["日期", "价格"]],
        column_widths={"A": 16, "B": 14},
    ),
    ProductConfig(
        name="Polymeric MDI",
        page_url="https://www.sci99.com/monitor-384-0.html",
        old_id="384",
        sheet_name="聚合MDI",
        preamble_rows=[["日期", "价格"]],
        column_widths={"A": 16, "B": 14},
    ),
    ProductConfig(
        name="电解铝",
        page_url="https://www.sci99.com/monitor-643-1.html",
        old_id="643",
        sheet_name="铝",
        preamble_rows=[[None, None], ["指标名称", "中国:平均价:铝(A00):有色市场"]],
        column_widths={"A": 16, "B": 18},
    ),
    ProductConfig(
        name="电解锡",
        page_url="https://www.sci99.com/monitor-647-1.html",
        old_id="647",
        sheet_name="锡",
        preamble_rows=[[None, None], ["指标名称", "中国:平均价:锡(1#):有色市场"]],
        column_widths={"A": 16, "B": 18},
    ),
]


def fetch_product_data(product: ProductConfig) -> list[dict[str, Any]]:
    response = requests.get(
        API_URL,
        params={"oldId": product.old_id, "type": "0"},
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": product.page_url,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=30,
    )
    response.raise_for_status()

    payload = response.json()
    if payload.get("code") != 200:
        raise RuntimeError(f"{product.name} 接口返回异常: {payload}")

    return payload.get("data") or []


def pick_workbook_path() -> Path:
    target = Path.cwd() / TARGET_FILE
    if target.exists():
        return target

    existing_excels = sorted(Path.cwd().glob("*.xlsx"))
    if existing_excels:
        return existing_excels[0]

    return target


def load_or_create_workbook(workbook_path: Path):
    if workbook_path.exists():
        return load_workbook(workbook_path)

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    return workbook


def normalize_trade_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text)
        except ValueError:
            return None
    return None


def to_float(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def build_r32_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                datetime.strptime(item["dateRange"], "%Y-%m-%d").date(),
                to_float(item["mdataValue"]),
                to_float(item["change"]),
                to_float(str(item["changeRate"]).rstrip("%")) / 100 if item.get("changeRate") not in {None, ""} else None,
                to_float(item["ndaysAvgPrice"]),
            ]
        )
    return rows


def build_pure_mdi_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    # 只留“日期 / 价格”两列；涨跌 / 幅度 / 七日均价这三列已移除
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                datetime.strptime(item["dateRange"], "%Y-%m-%d").date(),
                to_float(item["mdataValue"]),
            ]
        )
    return rows


def build_tdi_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                datetime.strptime(item["dateRange"], "%Y-%m-%d").date(),
                to_float(item["mdataValue"]),
            ]
        )
    return rows


def build_metal_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    # 铝/锡：仅保留“日期 / 价格”两列，丢弃涨跌/幅度/7日均
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                datetime.strptime(item["dateRange"], "%Y-%m-%d").date(),
                to_float(item["mdataValue"]),
            ]
        )
    return rows


def build_rows(product: ProductConfig, items: list[dict[str, Any]]) -> list[list[Any]]:
    if product.name == "R32":
        return build_r32_rows(items)
    if product.name in {"Pure MDI", "Polymeric MDI"}:
        return build_pure_mdi_rows(items)
    if product.name == "TDI":
        return build_tdi_rows(items)
    if product.name in {"电解铝", "电解锡"}:
        return build_metal_rows(items)
    raise ValueError(f"Unsupported product: {product.name}")


def read_existing_rows(workbook, product: ProductConfig) -> list[list[Any]]:
    if product.sheet_name not in workbook.sheetnames:
        return []

    worksheet = workbook[product.sheet_name]
    rows: list[list[Any]] = []
    data_start_row = len(product.preamble_rows) + 1

    for row_index in range(data_start_row, worksheet.max_row + 1):
        trade_date = normalize_trade_date(worksheet.cell(row=row_index, column=1).value)
        if trade_date is None:
            continue

        row_values = [trade_date]
        has_data = False
        for column_index in range(2, worksheet.max_column + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            row_values.append(value)
            if value not in {None, ""}:
                has_data = True

        if has_data:
            rows.append(row_values)

    return rows


def merge_rows(existing_rows: list[list[Any]], new_rows: list[list[Any]]) -> tuple[list[list[Any]], int, int]:
    rows_by_date: dict[date, list[Any]] = {row[0]: row for row in existing_rows}
    inserted = 0
    updated = 0

    for row in new_rows:
        trade_date = row[0]
        if trade_date in rows_by_date:
            updated += 1
        else:
            inserted += 1
        rows_by_date[trade_date] = row

    merged_rows = sorted(rows_by_date.values(), key=lambda item: item[0], reverse=True)
    return merged_rows, inserted, updated


def recreate_sheet(workbook, product: ProductConfig):
    if product.sheet_name in workbook.sheetnames:
        index = workbook.sheetnames.index(product.sheet_name)
        del workbook[product.sheet_name]
        return workbook.create_sheet(title=product.sheet_name, index=index)
    return workbook.create_sheet(title=product.sheet_name)


def write_preamble(worksheet, product: ProductConfig) -> None:
    for row in product.preamble_rows:
        worksheet.append(row)

    # R32 与 纯 MDI 同为「日期/价格/涨跌/幅度/七日均价」五列看板样式
    if product.name in {"R32", "Pure MDI", "Polymeric MDI"}:
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
        worksheet.row_dimensions[1].height = 24
        worksheet.freeze_panes = "A2"
    else:
        for row in worksheet.iter_rows(min_row=1, max_row=len(product.preamble_rows)):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = CENTER_ALIGNMENT
        worksheet.freeze_panes = "A3"

    for column_letter, width in product.column_widths.items():
        worksheet.column_dimensions[column_letter].width = width


def write_data_rows(worksheet, rows: list[list[Any]]) -> None:
    start_row = worksheet.max_row + 1
    for row in rows:
        worksheet.append(row)

    for row_index in range(start_row, worksheet.max_row + 1):
        for cell in worksheet[row_index]:
            cell.font = BODY_FONT
            cell.alignment = CENTER_ALIGNMENT


def apply_number_formats(worksheet, product: ProductConfig, rows: list[list[Any]]) -> None:
    start_row = len(product.preamble_rows) + 1
    for offset, row in enumerate(rows):
        row_index = start_row + offset
        worksheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"
        worksheet.cell(row=row_index, column=2).number_format = "0.00"
        if len(row) >= 3:
            worksheet.cell(row=row_index, column=3).number_format = "0.00"
        if len(row) >= 4:
            worksheet.cell(row=row_index, column=4).number_format = "0.00%"
        if len(row) >= 5:
            worksheet.cell(row=row_index, column=5).number_format = "0.00"


def save_workbook(workbook, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError:
        buffer = BytesIO()
        workbook.save(buffer)
        with workbook_path.open("rb+") as handle:
            handle.seek(0)
            handle.write(buffer.getvalue())
            handle.truncate()


def update_product_sheet(workbook, product: ProductConfig) -> tuple[int, int, int]:
    items = fetch_product_data(product)
    new_rows = build_rows(product, items)
    existing_rows = read_existing_rows(workbook, product)
    merged_rows, inserted, updated = merge_rows(existing_rows, new_rows)

    worksheet = recreate_sheet(workbook, product)
    write_preamble(worksheet, product)
    write_data_rows(worksheet, merged_rows)
    apply_number_formats(worksheet, product, merged_rows)

    return len(merged_rows), inserted, updated


def main() -> None:
    workbook_path = pick_workbook_path()
    workbook = load_or_create_workbook(workbook_path)

    summaries: list[str] = []
    for product in PRODUCTS:
        total_rows, inserted, updated = update_product_sheet(workbook, product)
        summaries.append(
            f"{product.sheet_name}: 共 {total_rows} 行，新增 {inserted} 行，覆盖 {updated} 行"
        )

    save_workbook(workbook, workbook_path)
    print(f"已更新工作簿: {workbook_path.name}")
    for summary in summaries:
        print(summary)


if __name__ == "__main__":
    main()
